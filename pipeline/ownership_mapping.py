"""
Ownership Mapping — Transparent per-company-year plant-to-company mapping.

Builds an explicit mapping of which GEM plants belong to each company,
including equity shares, and cross-checks against Kampmann's plant lists
to flag mismatches for human review.

Produces two outputs:
1. steel_ownership_mapping.csv — complete plant-company-year mapping
2. steel_ownership_mismatches.csv — flagged discrepancies for review

This module provides the transparency layer that Kampmann's methodology
requires: "Where available, we retrieved P(disclosed) and CAP(disclosed)
for production plants from mandatory financial reporting (e.g., annual
reports, investor presentations)..." — the ownership tree must be documented
and cross-referenced.

NOTE: GEM Plant IDs changed between the March 2023 vintage (Kampmann used
short codes like SFI00001) and December 2025 (now P100000120468). Cross-
referencing between the two datasets therefore uses **plant name + country
fuzzy matching**, not plant_id matching.
"""

import logging
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd

from .config import (
    GEM_STEEL_PLANTS_FILE,
    KAMPMANN_EXCEL_FILE,
    OWNERSHIP_MAPPING_FILE,
    OWNERSHIP_MISMATCHES_FILE,
    OUTPUTS_COMPANY_STEEL,
)
from .apa_calculator import (
    COMPANY_GEM_PATTERNS,
    COMPANY_PLANT_NAME_PATTERNS,
    OWNERSHIP_TRANSFERS,
    load_all_gem_plants,
    get_plants_for_year,
)

logger = logging.getLogger(__name__)


# ============================================================================
# Equity share parsing from GEM Parent field
# ============================================================================

def parse_parent_equity(parent_str: str) -> list[dict]:
    """Parse the GEM Parent field into (entity, equity_pct) pairs.

    GEM format examples:
      "ArcelorMittal SA [100.0%]"
      "ArcelorMittal SA [60.0%]; Nippon Steel Corp [40.0%]"
      "China Steel Corp; Formosa Plastics Corp; JFE Holdings Inc"
      "Tosyali Holding AS"

    Returns:
        List of dicts with keys: entity, equity_pct (0-1 or NaN if unknown)
    """
    if pd.isna(parent_str) or not str(parent_str).strip():
        return []

    parts = str(parent_str).split(";")
    result = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Extract equity percentage from [XX.X%]
        match = re.search(r"\[(\d+\.?\d*)\s*%\]", part)
        if match:
            pct = float(match.group(1)) / 100.0
            entity = re.sub(r"\s*\[\d+\.?\d*\s*%\]", "", part).strip()
        else:
            pct = np.nan  # Unknown equity
            entity = part.strip()
        result.append({"entity": entity, "equity_pct": pct})
    return result


def get_company_equity_share(parent_str: str, company_pattern: str) -> float:
    """Get the equity share for a specific company from the Parent field.

    Args:
        parent_str: GEM Parent field value
        company_pattern: Regex pattern matching the company

    Returns:
        Equity share (0-1), NaN if the company matches but no % given,
        or 0.0 if the company is not found in the Parent field.
    """
    parents = parse_parent_equity(parent_str)
    for p in parents:
        if re.search(company_pattern, p["entity"], re.IGNORECASE):
            if np.isnan(p["equity_pct"]):
                return np.nan
            return p["equity_pct"]
    return 0.0


# ============================================================================
# Plant name normalisation for cross-referencing
# ============================================================================

def _normalise_plant_name(name: str) -> str:
    """Normalise a plant name for fuzzy matching between GEM vintages.

    Strips common suffixes, company prefixes, and normalises whitespace/case
    so that "SSAB Raahe steel plant" matches "SSAB Raahe steel plant" across
    different GEM versions.
    """
    s = str(name).strip().lower()
    # Remove "steel plant", "steel works", etc. suffixes
    s = re.sub(r"\s+(steel|iron|works|plant|mill|steelworks|ironworks)\s*$", "", s)
    s = re.sub(r"\s+(steel|iron)\s+(plant|works|mill)\s*$", "", s)
    # Remove company name prefixes (common)
    for prefix in [
        "arcelormittal", "tata steel", "nippon steel", "posco", "ssab",
        "thyssenkrupp", "bluescope", "severstal", "baoshan", "nucor",
        "gerdau", "jfe", "jsw", "sail", "nlmk", "evraz", "liberty",
        "hyundai", "voestalpine", "salzgitter", "cleveland-cliffs",
        "am/ns", "us steel", "u.s. steel",
    ]:
        if s.startswith(prefix):
            s = s[len(prefix):].strip()
    # Normalise whitespace and special chars
    s = re.sub(r"[''`]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _match_plant_name(gem_name: str, gem_country: str,
                      kampmann_plants: pd.DataFrame) -> pd.Series | None:
    """Find a Kampmann plant matching a GEM plant by name+country.

    Uses normalised names. Falls back to substring matching if exact fails.

    Returns:
        The matching Kampmann row, or None if no match.
    """
    if kampmann_plants.empty:
        return None

    gem_norm = _normalise_plant_name(gem_name)
    gem_country_lower = str(gem_country).lower().strip()

    # First: exact normalised match + same country
    for _, kr in kampmann_plants.iterrows():
        k_norm = _normalise_plant_name(kr["plant_name"])
        k_country = str(kr["country"]).lower().strip()
        if gem_norm == k_norm and gem_country_lower == k_country:
            return kr

    # Second: one name contains the other + same country
    for _, kr in kampmann_plants.iterrows():
        k_norm = _normalise_plant_name(kr["plant_name"])
        k_country = str(kr["country"]).lower().strip()
        if gem_country_lower != k_country:
            continue
        if gem_norm in k_norm or k_norm in gem_norm:
            return kr

    # Third: key location word match + same country
    # Extract location words (non-generic words that identify the site)
    gem_words = set(gem_norm.split()) - {
        "steel", "plant", "works", "iron", "mill", "new", "old",
        "integrated", "facility", "complex", "base",
    }
    for _, kr in kampmann_plants.iterrows():
        k_norm = _normalise_plant_name(kr["plant_name"])
        k_country = str(kr["country"]).lower().strip()
        if gem_country_lower != k_country:
            continue
        k_words = set(k_norm.split()) - {
            "steel", "plant", "works", "iron", "mill", "new", "old",
            "integrated", "facility", "complex", "base",
        }
        if gem_words and k_words and gem_words & k_words:
            return kr

    return None


# ============================================================================
# Kampmann plant list extraction
# ============================================================================

def load_kampmann_plant_lists() -> pd.DataFrame:
    """Extract Kampmann's plant lists from his Excel workbook BAU sheets.

    Each BAU sheet has unit-level rows with:
    - Column B: GEM Plant ID (old format, e.g. SFI00001)
    - Column E: Plant name
    - Column F: Country
    - Column H: Ownership share (0-1)
    - Column I: Total nominal capacity (ttpa)
    - Column O: Main production process
    - Column Q: Status

    Returns:
        DataFrame with unique plants per company:
        company, kampmann_plant_id, plant_name, country, ownership_share,
        capacity_ttpa, process, status
    """
    if not KAMPMANN_EXCEL_FILE.exists():
        logger.warning(f"Kampmann Excel not found: {KAMPMANN_EXCEL_FILE}")
        return pd.DataFrame()

    import openpyxl
    wb = openpyxl.load_workbook(KAMPMANN_EXCEL_FILE, data_only=True, read_only=True)

    bau_sheets = [s for s in wb.sheetnames if s.strip().endswith("_BAU")]
    records = []

    _company_map = {
        "ArcelorMittal": "ArcelorMittal",
        "Baoshan": "Baoshan Iron & Steel",
        "Baoshan Iron & Steel": "Baoshan Iron & Steel",
        "BlueScope": "BlueScope Steel",
        "BlueScope Steel": "BlueScope Steel",
        "China Steel": "China Steel",
        "POSCO": "POSCO Holdings",
        "ThyssenKrupp": "ThyssenKrupp",
        "Severstal": "Severstal",
        "Nippon Steel": "Nippon Steel",
        "SSAB": "SSAB",
        "Tata Steel": "Tata Steel",
    }

    for sheet_name in bau_sheets:
        raw_name = sheet_name.strip().replace("_BAU_2", "").replace("_BAU", "").strip()
        company = _company_map.get(raw_name, raw_name)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        for row in rows[1:]:
            if len(row) < 17:
                continue

            plant_id = row[1]
            plant_name = row[4]
            country = row[5]
            ownership = row[7]
            total_cap = row[8]
            process = row[14]
            status = row[16]

            if not plant_id or not plant_name:
                continue
            if str(plant_id).strip().lower() in ("", "gem plant id", "nan"):
                continue
            if str(plant_id).strip().lower().startswith("dummy"):
                continue

            try:
                own_share = float(ownership) if ownership is not None else np.nan
            except (ValueError, TypeError):
                own_share = np.nan

            try:
                cap = float(total_cap) if total_cap is not None else 0.0
            except (ValueError, TypeError):
                cap = 0.0

            records.append({
                "company": company,
                "kampmann_plant_id": str(plant_id).strip() if plant_id else "",
                "plant_name": str(plant_name).strip() if plant_name else "",
                "country": str(country).strip() if country else "",
                "kampmann_ownership_share": own_share,
                "kampmann_capacity_ttpa": cap,
                "kampmann_process": str(process).strip() if process else "",
                "kampmann_status": str(status).strip() if status else "",
            })

    wb.close()

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # Deduplicate to unique plants per company (multiple units per plant)
    plant_level = (
        df.groupby(["company", "kampmann_plant_id"])
        .agg({
            "plant_name": "first",
            "country": "first",
            "kampmann_ownership_share": "first",
            "kampmann_capacity_ttpa": "sum",
            "kampmann_process": lambda x: "; ".join(sorted(set(
                str(v) for v in x if str(v) not in ("", "nan", "None")
            ))),
            "kampmann_status": "first",
        })
        .reset_index()
    )

    # Filter out transition/dummy entries (IDs with "-1", "-2" suffix = technology transitions)
    plant_level = plant_level[
        ~plant_level["kampmann_plant_id"].str.contains(r"-\d+$", regex=True)
    ].copy()

    logger.info(f"Kampmann plant lists: {len(plant_level)} unique plants "
                f"across {plant_level['company'].nunique()} companies")
    return plant_level


# ============================================================================
# Ownership mapping generation
# ============================================================================

def generate_ownership_mapping(years: list[int] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Generate comprehensive ownership mapping with cross-checks.

    For each company and year:
    1. Extract GEM plants matched to the company (with equity shares)
    2. Cross-reference with Kampmann's plant list by NAME+COUNTRY
       (GEM Plant IDs changed between Mar 2023 and Dec 2025 vintages)
    3. Flag mismatches: plants in GEM but not Kampmann, or vice versa
    4. Flag equity share differences
    5. Flag minority stakes that may affect consolidation boundary

    Args:
        years: Years to generate mapping for. Defaults to 2014-2024.

    Returns:
        Tuple of (mapping_df, mismatches_df)
    """
    if years is None:
        years = list(range(2014, 2025))

    # Load GEM plants
    all_plants = load_all_gem_plants()
    if all_plants.empty:
        logger.error("No GEM plants loaded")
        return pd.DataFrame(), pd.DataFrame()

    # Load Kampmann plant lists for cross-checking
    kampmann_plants = load_kampmann_plant_lists()
    has_kampmann = not kampmann_plants.empty

    # Load raw GEM plant data for full Parent field (with equity info)
    gem_raw = pd.read_excel(GEM_STEEL_PLANTS_FILE, sheet_name="Plant data")
    parent_col = [c for c in gem_raw.columns if "parent" in c.lower()][0]
    plant_id_col = [c for c in gem_raw.columns if "plant id" in c.lower()][0]
    raw_parent_map = dict(zip(
        gem_raw[plant_id_col].astype(str),
        gem_raw[parent_col].astype(str),
    ))

    mapping_rows = []
    mismatch_rows = []

    for company, pattern in COMPANY_GEM_PATTERNS.items():
        # Get Kampmann plants for this company (for cross-checking)
        if has_kampmann:
            k_plants_co = kampmann_plants[kampmann_plants["company"] == company]
        else:
            k_plants_co = pd.DataFrame()

        # We only need to generate the mapping once per year (not repeat per
        # Kampmann year). But we DO cross-check only at 2020 (Kampmann's base
        # year) to avoid spurious mismatches from year-specific plant filtering.
        for year in years:
            year_plants = get_plants_for_year(all_plants, year)
            company_plants = _get_company_plants_with_equity(
                year_plants, company, pattern, year, raw_parent_map
            )

            if company_plants.empty:
                continue

            # Build set of matched GEM plant names (normalised) for cross-check
            gem_names_matched = set()

            for _, plant in company_plants.iterrows():
                pid = str(plant["plant_id"]).strip()
                parent_raw = raw_parent_map.get(pid, "")
                equity = plant.get("equity_share", np.nan)
                plant_name = str(plant.get("plant_name", ""))
                country = str(plant.get("country", ""))

                # Cross-check with Kampmann by plant name
                k_match = None
                in_kampmann = False
                k_ownership = np.nan
                k_process = ""
                k_plant_id = ""
                if not k_plants_co.empty:
                    k_match = _match_plant_name(plant_name, country, k_plants_co)
                    if k_match is not None:
                        in_kampmann = True
                        k_ownership = k_match.get("kampmann_ownership_share", np.nan)
                        k_process = k_match.get("kampmann_process", "")
                        k_plant_id = k_match.get("kampmann_plant_id", "")
                        gem_names_matched.add(
                            _normalise_plant_name(str(k_match.get("plant_name", "")))
                        )

                # Determine flags
                flags = []
                is_kampmann_company = company in set(kampmann_plants["company"]) if has_kampmann else False

                if not in_kampmann and is_kampmann_company:
                    flags.append("NOT_IN_KAMPMANN")

                if in_kampmann:
                    # Check equity mismatch
                    if pd.notna(equity) and pd.notna(k_ownership):
                        if abs(equity - k_ownership) > 0.02:
                            flags.append(
                                f"EQUITY_MISMATCH_K={k_ownership:.0%}_GEM={equity:.0%}"
                            )

                if pd.isna(equity):
                    flags.append("EQUITY_UNKNOWN")

                if pd.notna(equity) and equity < 0.5:
                    flags.append(f"MINORITY_STAKE_{equity:.0%}")

                mapping_rows.append({
                    "company": company,
                    "year": year,
                    "plant_id": pid,
                    "plant_name": plant_name,
                    "country": country,
                    "status": plant.get("status", ""),
                    "capacity_ttpa": plant.get("capacity_ttpa", 0),
                    "process": plant.get("process", ""),
                    "gem_parent_raw": parent_raw,
                    "equity_share": equity,
                    "kampmann_plant_id": k_plant_id,
                    "kampmann_ownership_share": k_ownership if in_kampmann else np.nan,
                    "kampmann_process": k_process if in_kampmann else "",
                    "in_kampmann": in_kampmann,
                    "match_source": plant.get("match_source", "parent_pattern"),
                    "flags": "; ".join(flags) if flags else "",
                })

                # Record mismatches for review
                if flags:
                    for flag in flags:
                        flag_type = flag.split("_")[0] if "_" in flag else flag
                        if flag.startswith("EQUITY_MISMATCH"):
                            flag_type = "EQUITY_MISMATCH"
                        elif flag.startswith("MINORITY_STAKE"):
                            flag_type = "MINORITY_STAKE"
                        elif flag.startswith("NOT_IN"):
                            flag_type = "NOT_IN_KAMPMANN"

                        mismatch_rows.append({
                            "company": company,
                            "year": year,
                            "plant_id": pid,
                            "plant_name": plant_name,
                            "country": country,
                            "capacity_ttpa": plant.get("capacity_ttpa", 0),
                            "flag_type": flag_type,
                            "flag_detail": flag,
                            "gem_equity": equity,
                            "kampmann_equity": k_ownership if in_kampmann else np.nan,
                            "gem_parent_raw": parent_raw,
                            "action_needed": _suggest_action(flag, company, plant),
                        })

            # Check for Kampmann plants NOT found in GEM (at year 2020 only,
            # to avoid spurious mismatches from year-specific filtering)
            if year == 2020 and not k_plants_co.empty:
                for _, kr in k_plants_co.iterrows():
                    k_norm = _normalise_plant_name(str(kr["plant_name"]))
                    k_status = str(kr.get("kampmann_status", "")).lower()
                    if k_status in ("cancelled", "announced"):
                        continue
                    if k_norm not in gem_names_matched:
                        mismatch_rows.append({
                            "company": company,
                            "year": year,
                            "plant_id": kr.get("kampmann_plant_id", ""),
                            "plant_name": kr.get("plant_name", ""),
                            "country": kr.get("country", ""),
                            "capacity_ttpa": kr.get("kampmann_capacity_ttpa", 0),
                            "flag_type": "IN_KAMPMANN_NOT_GEM",
                            "flag_detail": (
                                f"Plant '{kr.get('plant_name', '')}' in Kampmann "
                                f"({k_status}) not matched in GEM Dec 2025 for "
                                f"{company}. May be GEM vintage difference or "
                                f"name change."
                            ),
                            "gem_equity": np.nan,
                            "kampmann_equity": kr.get("kampmann_ownership_share", np.nan),
                            "gem_parent_raw": "",
                            "action_needed": (
                                "Check if this plant exists in GEM Dec 2025 under "
                                "a different name or parent. May require updating "
                                "COMPANY_GEM_PATTERNS or manual name mapping."
                            ),
                        })

    mapping_df = pd.DataFrame(mapping_rows)
    mismatches_df = pd.DataFrame(mismatch_rows)

    if not mapping_df.empty:
        mapping_df = mapping_df.sort_values(
            ["company", "year", "country", "plant_name"]
        ).reset_index(drop=True)

    if not mismatches_df.empty:
        mismatches_df = mismatches_df.sort_values(
            ["company", "year", "flag_type", "plant_name"]
        ).reset_index(drop=True)

    # Save outputs
    OUTPUTS_COMPANY_STEEL.mkdir(parents=True, exist_ok=True)

    mapping_df.to_csv(OWNERSHIP_MAPPING_FILE, index=False)
    logger.info(f"Saved ownership mapping: {OWNERSHIP_MAPPING_FILE} "
                f"({len(mapping_df)} rows, "
                f"{mapping_df['company'].nunique() if not mapping_df.empty else 0} companies)")

    mismatches_df.to_csv(OWNERSHIP_MISMATCHES_FILE, index=False)
    logger.info(f"Saved ownership mismatches: {OWNERSHIP_MISMATCHES_FILE} "
                f"({len(mismatches_df)} rows)")

    # Print summary
    _print_summary(mapping_df, mismatches_df)

    return mapping_df, mismatches_df


def _get_company_plants_with_equity(
    plants_df: pd.DataFrame,
    company: str,
    pattern: str,
    year: int,
    raw_parent_map: dict[str, str],
) -> pd.DataFrame:
    """Get company plants with equity share extracted from raw GEM Parent field.

    Enhances get_company_plants() by also extracting the equity share
    from the full GEM Parent field (which contains [XX.X%] annotations).
    """
    mask = plants_df["parent"].str.contains(pattern, case=False, na=False)
    result = plants_df[mask].copy()
    result["match_source"] = "parent_pattern"

    # Handle ownership transfers
    if not result.empty:
        for transfer in OWNERSHIP_TRANSFERS:
            if transfer["acquirer"] == company and year < transfer["year_acquired"]:
                exclude_pat = transfer["target_plant_pattern"]
                exclude_mask = result["plant_name"].str.contains(
                    exclude_pat, case=False, na=False
                )
                if exclude_mask.any():
                    result = result[~exclude_mask].copy()

    # Fallback: match by plant name
    if result.empty and company in COMPANY_PLANT_NAME_PATTERNS:
        name_pat = COMPANY_PLANT_NAME_PATTERNS[company]
        name_mask = plants_df["plant_name"].str.contains(name_pat, case=False, na=False)
        result = plants_df[name_mask].copy()
        result["match_source"] = "plant_name_fallback"

    if result.empty:
        return result

    # Extract equity share from raw Parent field
    equity_shares = []
    for _, row in result.iterrows():
        pid = str(row["plant_id"])
        parent_raw = raw_parent_map.get(pid, "")
        equity = get_company_equity_share(parent_raw, pattern)
        equity_shares.append(equity)
    result["equity_share"] = equity_shares

    return result


def _suggest_action(flag: str, company: str, plant) -> str:
    """Suggest a review action for a flagged mismatch."""
    pname = plant.get("plant_name", "") if hasattr(plant, "get") else ""
    if flag.startswith("NOT_IN_KAMPMANN"):
        return (f"Check if '{pname}' is consolidated in "
                f"{company}'s annual report. May be GEM vintage difference "
                f"(Dec 2025 vs Mar 2023) or new plant since Kampmann's analysis.")
    if flag.startswith("EQUITY_MISMATCH"):
        return ("Verify equity share from annual report. GEM Dec 2025 and "
                "Kampmann Mar 2023 may use different reporting dates.")
    if flag.startswith("EQUITY_UNKNOWN"):
        return ("GEM Parent field has no [XX.X%] for this company. Check "
                "annual report for actual ownership percentage.")
    if flag.startswith("MINORITY_STAKE"):
        return (f"Equity stake below 50%. Check if {company} consolidates "
                f"this plant (full consolidation) or only equity-share production.")
    return "Review needed"


def _print_summary(mapping: pd.DataFrame, mismatches: pd.DataFrame):
    """Print a summary of the ownership mapping."""
    print("\n" + "=" * 80)
    print("OWNERSHIP MAPPING SUMMARY")
    print("=" * 80)

    if mapping.empty:
        print("No mapping data generated.")
        return

    # Per-company summary at 2020 (base year)
    m2020 = mapping[mapping["year"] == 2020]
    print(f"\nBase year 2020: {len(m2020)} plant-entries across "
          f"{m2020['company'].nunique()} companies")
    print(f"\n{'Company':<25} {'Plants':>6} {'Cap(Mt)':>8} "
          f"{'100%':>5} {'Part':>5} {'Unk':>5} {'InK':>5} {'Flags':>6}")
    print("-" * 80)

    for company in sorted(m2020["company"].unique()):
        c = m2020[m2020["company"] == company]
        n_plants = c["plant_id"].nunique()
        total_cap = c["capacity_ttpa"].sum() / 1000.0
        n_full = ((c["equity_share"] > 0.99) & c["equity_share"].notna()).sum()
        n_partial = ((c["equity_share"] <= 0.99) & (c["equity_share"] > 0) & c["equity_share"].notna()).sum()
        n_unknown = c["equity_share"].isna().sum()
        n_in_k = c["in_kampmann"].sum()
        n_flags = (c["flags"] != "").sum()
        print(f"{company:<25} {n_plants:>6} {total_cap:>8.1f} "
              f"{n_full:>5} {n_partial:>5} {n_unknown:>5} {n_in_k:>5} {n_flags:>6}")

    # Mismatch summary
    if not mismatches.empty:
        # Deduplicate: only show unique (company, plant_name, flag_type) combos
        unique_mm = mismatches.drop_duplicates(
            subset=["company", "plant_name", "flag_type"]
        )
        print(f"\n--- MISMATCHES: {len(unique_mm)} unique flags ---")
        flag_counts = unique_mm["flag_type"].value_counts()
        for flag_type, count in flag_counts.items():
            print(f"  {flag_type}: {count}")

        print(f"\n{'Company':<25} {'Plant':<40} {'Country':<12} {'Flag':<22} {'Action'}")
        print("-" * 140)
        for _, m in unique_mm.head(40).iterrows():
            plant_short = str(m["plant_name"])[:38]
            country_short = str(m["country"])[:10]
            flag_short = str(m["flag_type"])[:20]
            action_short = str(m["action_needed"])[:55]
            print(f"{m['company']:<25} {plant_short:<40} {country_short:<12} "
                  f"{flag_short:<22} {action_short}")
        if len(unique_mm) > 40:
            print(f"  ... and {len(unique_mm) - 40} more (see CSV for full list)")

    print("=" * 80)


# ============================================================================
# Standalone execution
# ============================================================================

def main():
    """Run ownership mapping standalone."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )

    print("=" * 60)
    print("STEEL OWNERSHIP MAPPING GENERATOR")
    print("=" * 60)

    mapping_df, mismatches_df = generate_ownership_mapping()

    if not mapping_df.empty:
        print(f"\nMapping: {len(mapping_df)} rows saved to {OWNERSHIP_MAPPING_FILE}")
    if not mismatches_df.empty:
        print(f"Mismatches: {len(mismatches_df)} rows saved to {OWNERSHIP_MISMATCHES_FILE}")


if __name__ == "__main__":
    main()
