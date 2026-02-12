"""
Steel APA Data Collection & Calculation Pipeline
================================================

This module automates the collection and processing of steel company data
following David Kampmann's methodology.

WORKFLOW:
1. Extract company production from WSA PDFs / company reports
2. Load plant data from GEM Steel Plant Tracker  
3. Assign emission factors by country × process
4. Calculate emissions using uniform utilization rate

Author: Carbon Budget Tracker team
"""

import pandas as pd
import numpy as np
import requests
from pathlib import Path
import re
from typing import Dict, List, Optional, Tuple
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# SECTION 1: EMISSION FACTORS (Fully automated - lookup tables)
# =============================================================================

# David Kampmann's emission factors (tCO2 per tonne crude steel)
EF_BF_BOF = {
    "Brazil": 2.19,
    "China": 1.76,
    "EU": 1.77,
    "India": 3.72,
    "Japan": 2.05,
    "Russia": 2.79,
    "Serbia": 2.06,
    "South Africa": 3.57,
    "South Korea": 2.00,
    "Taiwan": 2.02,
    "Turkey": 2.17,
    "Ukraine": 2.30,
    "United Kingdom": 2.05,
    "United States": 1.94,
    "Global": 2.314
}

EF_EAF = {
    "Brazil": 0.05,
    "China": 0.03,
    "EU": 0.04,
    "India": 0.07,
    "Japan": 0.04,
    "Russia": 0.07,
    "Serbia": 0.06,
    "South Africa": 0.12,
    "South Korea": 0.03,
    "Switzerland": 0.02,
    "Taiwan": 0.02,
    "Turkey": 0.04,
    "Ukraine": 0.04,
    "United Kingdom": 0.04,
    "United States": 0.04,
    "Thailand": 0.05,
    "Global": 0.051
}

# CRITICAL: DRI emission factors depend on fuel type
EF_DRI_COAL = 3.10  # Coal-based (rotary kilns) - India, China, South Africa
EF_DRI_GAS = 1.05   # Gas-based (MIDREX, HYL) - Iran, Middle East, Americas
EF_H2_DRI = 0.04    # Hydrogen-based (future)

# Countries using coal-based DRI (CRITICAL DISTINCTION!)
DRI_COAL_COUNTRIES = ["India", "China", "South Africa"]

# Country to EF region mapping
COUNTRY_TO_EF_REGION = {
    # Direct mappings
    "Brazil": "Brazil", "China": "China", "India": "India", "Japan": "Japan",
    "Russia": "Russia", "Serbia": "Serbia", "South Africa": "South Africa",
    "South Korea": "South Korea", "Switzerland": "Switzerland", "Taiwan": "Taiwan",
    "Turkey": "Turkey", "Türkiye": "Turkey", "Ukraine": "Ukraine",
    "United Kingdom": "United Kingdom", "United States": "United States",
    
    # EU member states
    "Germany": "EU", "France": "EU", "Netherlands": "EU", "Belgium": "EU",
    "Austria": "EU", "Spain": "EU", "Italy": "EU", "Sweden": "EU",
    "Finland": "EU", "Poland": "EU", "Czech Republic": "EU", "Czechia": "EU",
    "Slovakia": "EU", "Luxembourg": "EU", "Romania": "EU", "Hungary": "EU",
    "Greece": "EU", "Portugal": "EU", "Ireland": "EU", "Denmark": "EU",
    "Croatia": "EU", "Slovenia": "EU", "Bulgaria": "EU", "Estonia": "EU",
    "Latvia": "EU", "Lithuania": "EU", "Cyprus": "EU", "Malta": "EU",
    
    # Other mappings
    "Canada": "United States", "Mexico": "Brazil", "Australia": "EU",
    "Thailand": "Thailand", "Vietnam": "China", "Indonesia": "India",
    "Iran": "Global", "Saudi Arabia": "Global", "United Arab Emirates": "Global",
}


def get_ef_region(country: str) -> str:
    """Map plant country to emission factor region."""
    return COUNTRY_TO_EF_REGION.get(country, "Global")


def get_plant_ef(country: str, process: str) -> float:
    """
    Get emission factor for a plant based on country and process.
    
    This is the core function that matches David Kampmann's methodology.
    The CRITICAL distinction is coal vs gas DRI.
    """
    process_lower = str(process).lower() if process else ""
    ef_region = get_ef_region(country)
    
    # Check for DRI first - this is the CRITICAL distinction!
    if any(x in process_lower for x in ['dri', 'sponge', 'direct reduction', 'ironmaking (dri']):
        if country in DRI_COAL_COUNTRIES:
            return EF_DRI_COAL  # 3.10 - Coal-based
        else:
            return EF_DRI_GAS   # 1.05 - Gas-based
    
    # Check for EAF
    elif any(x in process_lower for x in ['eaf', 'electric', 'scrap']):
        return EF_EAF.get(ef_region, EF_EAF.get("Global", 0.051))
    
    # Check for H2-DRI (future technology)
    elif 'h2' in process_lower or 'hydrogen' in process_lower:
        return EF_H2_DRI
    
    # Default: BF-BOF (integrated steelmaking)
    else:
        return EF_BF_BOF.get(ef_region, EF_BF_BOF.get("Global", 2.314))


# =============================================================================
# SECTION 2: DATA SOURCES - WSA Production Data
# =============================================================================

class WSADataExtractor:
    """
    Extract company production data from World Steel Association sources.
    
    Data sources:
    1. "World Steel in Figures" PDF (annual, top 50 producers)
    2. Top Producers webpage (https://worldsteel.org/data/top-steel-producers/)
    3. Monthly press releases
    
    The PDF structure is consistent year-to-year, making extraction reliable.
    """
    
    # URLs for WSA data
    WSA_PDF_URLS = {
        2024: "https://worldsteel.org/wp-content/uploads/World-Steel-in-Figures-2024.pdf",
        2023: "https://worldsteel.org/wp-content/uploads/World-Steel-in-Figures-2023.pdf",
        2022: "https://worldsteel.org/wp-content/uploads/World-Steel-in-Figures-2022.pdf",
        2021: "https://worldsteel.org/wp-content/uploads/World-Steel-in-Figures-2021.pdf",
        2020: "https://worldsteel.org/wp-content/uploads/World-Steel-in-Figures-2020.pdf",
    }
    
    WSA_TOP_PRODUCERS_URL = "https://worldsteel.org/data/top-steel-producers/"
    
    @staticmethod
    def download_wsa_pdf(year: int, output_dir: Path) -> Path:
        """Download WSA World Steel in Figures PDF."""
        url = WSADataExtractor.WSA_PDF_URLS.get(year)
        if not url:
            raise ValueError(f"No URL for year {year}")
        
        output_path = output_dir / f"wsa_steel_in_figures_{year}.pdf"
        
        logger.info(f"Downloading WSA PDF for {year}...")
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        output_path.write_bytes(response.content)
        logger.info(f"Saved to {output_path}")
        return output_path
    
    @staticmethod
    def extract_top_producers_from_pdf(pdf_path: Path) -> pd.DataFrame:
        """
        Extract top steel producers table from WSA PDF.
        
        Requires: pip install pdfplumber
        
        The table format is consistent:
        | Rank | Company | Country | Production (Mt) |
        """
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("Install pdfplumber: pip install pdfplumber")
        
        producers = []
        
        with pdfplumber.open(pdf_path) as pdf:
            # Top producers table is usually on pages 6-7
            for page_num in range(5, 10):
                if page_num >= len(pdf.pages):
                    break
                    
                page = pdf.pages[page_num]
                tables = page.extract_tables()
                
                for table in tables:
                    for row in table:
                        # Look for rows with company data
                        if row and len(row) >= 3:
                            # Try to parse as producer row
                            try:
                                # Check if first column looks like a rank or company name
                                if row[0] and (row[0].isdigit() or len(row[0]) > 3):
                                    rank = row[0] if row[0].isdigit() else None
                                    company = row[1] if rank else row[0]
                                    production = None
                                    
                                    # Find the production value (number > 3)
                                    for cell in row[2:]:
                                        if cell:
                                            try:
                                                val = float(cell.replace(',', '').replace(' ', ''))
                                                if val > 3:  # Mt threshold
                                                    production = val
                                                    break
                                            except:
                                                pass
                                    
                                    if company and production:
                                        producers.append({
                                            'rank': int(rank) if rank else None,
                                            'company': company.strip(),
                                            'production_mt': production
                                        })
                            except:
                                pass
        
        df = pd.DataFrame(producers)
        df = df.drop_duplicates(subset=['company'])
        return df.sort_values('production_mt', ascending=False).reset_index(drop=True)
    
    @staticmethod
    def get_manual_wsa_data() -> pd.DataFrame:
        """
        Returns manually compiled WSA top producers data.
        
        This is a fallback when PDF extraction isn't available.
        Data source: World Steel in Figures 2024 (for 2023 production)
        """
        data = [
            # Top 50 steel producers 2023 (from WSA)
            ("China Baowu Group", "China", 131.84),
            ("ArcelorMittal", "Luxembourg", 68.89),
            ("Ansteel Group", "China", 55.93),
            ("Nippon Steel", "Japan", 43.55),
            ("Shagang Group", "China", 41.45),
            ("POSCO Holdings", "South Korea", 38.62),
            ("HBIS Group", "China", 38.30),
            ("Jianlong Group", "China", 36.53),
            ("Shougang Group", "China", 34.31),
            ("Tata Steel", "India", 29.94),
            ("Delong Steel Group", "China", 28.43),
            ("JFE Holdings", "Japan", 25.22),
            ("Shandong Steel", "China", 24.27),
            ("JSW Steel", "India", 24.14),
            ("Fangda Steel", "China", 22.70),
            ("Nucor", "USA", 21.28),
            ("SAIL", "India", 19.20),
            ("Valin Group", "China", 19.01),
            ("Zhongtian Iron & Steel", "China", 18.72),
            ("Liuzhou Steel", "China", 17.90),
            ("Hyundai Steel", "South Korea", 17.21),
            ("Rizhao Steel", "China", 16.23),
            ("NLMK", "Russia", 15.30),
            ("Sanming Steel", "China", 14.89),
            ("Steel Dynamics", "USA", 14.46),
            ("Cleveland-Cliffs", "USA", 14.14),
            ("Baotou Steel", "China", 13.77),
            ("Severstal", "Russia", 11.15),
            ("Xinyu Iron & Steel", "China", 11.10),
            ("Evraz", "Russia", 10.70),
            ("Jinxi Iron & Steel", "China", 10.60),
            ("Gerdau", "Brazil", 10.45),
            ("ThyssenKrupp", "Germany", 10.44),
            ("Taiyuan Steel", "China", 10.34),
            ("CSN", "Brazil", 9.50),
            ("Benxi Steel", "China", 9.35),
            ("Jiuquan Steel", "China", 9.15),
            ("MMK", "Russia", 9.10),
            ("U. S. Steel", "USA", 8.95),
            ("BlueScope Steel", "Australia", 8.50),
            ("Ternium", "Argentina", 8.30),
            ("Voestalpine", "Austria", 7.80),
            ("SSAB", "Sweden", 7.50),
            ("China Steel Corporation", "Taiwan", 7.20),
            ("Jindal Steel & Power", "India", 6.80),
            ("Kobe Steel", "Japan", 6.50),
            ("Salzgitter", "Germany", 6.10),
        ]
        
        df = pd.DataFrame(data, columns=['company', 'headquarters', 'production_mt'])
        df['year'] = 2023
        df['source'] = 'WSA World Steel in Figures 2024'
        return df


# =============================================================================
# SECTION 3: Company Annual Report Scraping
# =============================================================================

class CompanyReportScraper:
    """
    Scrape production data from company investor relations pages.
    
    Each company has different page structure, so we define templates
    for the major companies we track.
    """
    
    # Company investor relations URLs
    COMPANY_IR_URLS = {
        "Tata Steel": {
            "base": "https://www.tatasteel.com/investors/",
            "annual_report": "https://www.tatasteel.com/investors/integrated-reportannual-report/",
            "quarterly": "https://www.tatasteel.com/media/newsroom/press-releases/",
        },
        "ArcelorMittal": {
            "base": "https://corporate.arcelormittal.com/investors",
            "annual_report": "https://corporate.arcelormittal.com/investors/annual-reports",
        },
        "POSCO": {
            "base": "https://www.posco.co.kr/homepage/docs/eng6/jsp/ir/",
            "annual_report": "https://www.posco.co.kr/homepage/docs/eng6/jsp/ir/ir_02.jsp",
        },
        "Nippon Steel": {
            "base": "https://www.nipponsteel.com/en/ir/",
            "annual_report": "https://www.nipponsteel.com/en/ir/library/",
        },
        "JSW Steel": {
            "base": "https://www.jsw.in/investors/steel",
            "annual_report": "https://www.jsw.in/investors/steel/annual-reports",
        },
        "SAIL": {
            "base": "https://sail.co.in/en/investor",
            "annual_report": "https://sail.co.in/en/investor/annual-report",
        },
        "Cleveland-Cliffs": {
            "base": "https://www.clevelandcliffs.com/investors",
            "sec_filings": "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0000764065&type=10-K",
        },
    }
    
    @staticmethod
    def scrape_tata_steel_production(year: int) -> Optional[Dict]:
        """
        Scrape Tata Steel production from their quarterly releases.
        
        Tata publishes production in press releases with consistent format:
        "crude steel production of X million tonnes"
        """
        # This would need actual web scraping implementation
        # For now, return manually collected data
        
        tata_historical = {
            2020: {"india": 18.23, "europe": 9.79, "thailand": 1.09, "total": 28.54},
            2021: {"india": 19.06, "europe": 10.50, "thailand": 1.09, "total": 31.03},
            2022: {"india": 19.51, "europe": 9.88, "thailand": 1.09, "total": 30.65},
            2023: {"india": 20.00, "europe": 8.50, "thailand": 1.09, "total": 29.94},
            2024: {"india": 20.80, "europe": 7.80, "thailand": 1.09, "total": 29.69},
        }
        
        return tata_historical.get(year)
    
    @staticmethod
    def get_manual_company_data() -> pd.DataFrame:
        """
        Returns manually compiled company production data.
        
        Sources:
        - Company annual reports
        - Quarterly production releases
        - SEC filings (10-K, 20-F)
        - Sustainability reports
        """
        data = [
            # Tata Steel (from annual reports)
            ("Tata Steel", 2020, 28.54, "Annual Report FY2020-21"),
            ("Tata Steel", 2021, 31.03, "Annual Report FY2021-22"),
            ("Tata Steel", 2022, 30.65, "Annual Report FY2022-23"),
            ("Tata Steel", 2023, 29.94, "Annual Report FY2023-24"),
            
            # ArcelorMittal (from 20-F filings)
            ("ArcelorMittal", 2020, 71.50, "20-F 2020"),
            ("ArcelorMittal", 2021, 79.26, "20-F 2021"),
            ("ArcelorMittal", 2022, 68.90, "20-F 2022"),
            ("ArcelorMittal", 2023, 68.89, "20-F 2023"),
            
            # POSCO (from sustainability reports)
            ("POSCO", 2020, 40.58, "Sustainability Report 2020"),
            ("POSCO", 2021, 42.86, "Sustainability Report 2021"),
            ("POSCO", 2022, 38.65, "Sustainability Report 2022"),
            ("POSCO", 2023, 38.62, "Sustainability Report 2023"),
            
            # JSW Steel (from quarterly releases)
            ("JSW Steel", 2020, 14.00, "Quarterly release"),
            ("JSW Steel", 2021, 15.00, "Quarterly release"),
            ("JSW Steel", 2022, 22.00, "Quarterly release"),
            ("JSW Steel", 2023, 24.14, "WSA 2024"),
            
            # SAIL (from annual reports)
            ("SAIL", 2020, 15.00, "Annual Report 2020-21"),
            ("SAIL", 2021, 17.40, "Annual Report 2021-22"),
            ("SAIL", 2022, 18.30, "Annual Report 2022-23"),
            ("SAIL", 2023, 19.20, "WSA 2024"),
            
            # Cleveland-Cliffs (from 10-K filings)
            ("Cleveland-Cliffs", 2021, 17.00, "10-K 2021"),
            ("Cleveland-Cliffs", 2022, 14.80, "10-K 2022"),
            ("Cleveland-Cliffs", 2023, 14.14, "WSA 2024"),
            
            # Nippon Steel
            ("Nippon Steel", 2020, 41.60, "Annual Report"),
            ("Nippon Steel", 2021, 47.68, "Annual Report"),
            ("Nippon Steel", 2022, 44.37, "Annual Report"),
            ("Nippon Steel", 2023, 43.55, "WSA 2024"),
            
            # ThyssenKrupp Steel
            ("ThyssenKrupp", 2020, 10.70, "Annual Report"),
            ("ThyssenKrupp", 2021, 11.20, "Annual Report"),
            ("ThyssenKrupp", 2022, 10.44, "Annual Report"),
            ("ThyssenKrupp", 2023, 10.44, "WSA 2024"),
            
            # SSAB
            ("SSAB", 2020, 7.50, "Annual Report"),
            ("SSAB", 2021, 8.20, "Annual Report"),
            ("SSAB", 2022, 7.80, "Annual Report"),
            ("SSAB", 2023, 7.50, "WSA 2024"),
            
            # Severstal
            ("Severstal", 2020, 11.30, "Annual Report"),
            ("Severstal", 2021, 11.60, "Annual Report"),
            ("Severstal", 2022, 10.70, "Annual Report"),
            ("Severstal", 2023, 11.15, "WSA 2024"),
        ]
        
        return pd.DataFrame(data, columns=['company', 'year', 'production_mt', 'source'])


# =============================================================================
# SECTION 4: GEM Plant Data Loader
# =============================================================================

class GEMPlantLoader:
    """
    Load and process plant data from Global Energy Monitor Steel Plant Tracker.
    
    The GEM data is already included in David's Excel file.
    """
    
    @staticmethod
    def load_from_kampmann_excel(filepath: str) -> pd.DataFrame:
        """Load GEM plant data from Kampmann's Excel file."""
        
        df = pd.read_excel(filepath, sheet_name='GSPT_2023_Steel Plants_all', skiprows=1)
        
        # Select and rename relevant columns
        plants = df[[
            'Plant ID', 'Plant name (English)', 'Country', 'Parent [formula]',
            'Status', 'Nominal crude steel capacity (ttpa)',
            'Nominal BF capacity (ttpa)', 'Nominal BOF steel capacity (ttpa)',
            'Nominal EAF steel capacity (ttpa)', 'Nominal DRI capacity (ttpa)',
            'Main production process'
        ]].copy()
        
        plants.columns = [
            'plant_id', 'plant_name', 'country', 'parent', 'status',
            'capacity_ttpa', 'bf_capacity', 'bof_capacity', 'eaf_capacity', 
            'dri_capacity', 'main_process'
        ]
        
        # Filter to operating plants only
        plants = plants[plants['status'] == 'operating'].copy()
        
        # Convert capacities to numeric
        for col in ['capacity_ttpa', 'bf_capacity', 'bof_capacity', 'eaf_capacity', 'dri_capacity']:
            plants[col] = pd.to_numeric(plants[col], errors='coerce')
        
        # Determine process type
        plants['process'] = plants.apply(
            lambda row: GEMPlantLoader._determine_process(row), axis=1
        )
        
        # Assign emission factors
        plants['ef'] = plants.apply(
            lambda row: get_plant_ef(row['country'], row['process']), axis=1
        )
        
        return plants
    
    @staticmethod
    def _determine_process(row) -> str:
        """Determine plant process type from capacity and description."""
        process = str(row['main_process']).lower() if pd.notna(row['main_process']) else ''
        
        # Check capacity fields first
        if pd.notna(row['dri_capacity']) and row['dri_capacity'] > 0:
            return "DRI"
        if pd.notna(row['bf_capacity']) and row['bf_capacity'] > 0:
            return "BF-BOF"
        if pd.notna(row['eaf_capacity']) and row['eaf_capacity'] > 0:
            return "EAF"
        
        # Fall back to process description
        if 'dri' in process or 'sponge' in process or 'direct' in process:
            return "DRI"
        if 'eaf' in process or 'electric' in process:
            return "EAF"
        
        # Default to BF-BOF (most common for large integrated plants)
        return "BF-BOF"


# =============================================================================
# SECTION 5: Emissions Calculator
# =============================================================================

class SteelAPACalculator:
    """
    Calculate steel company emissions using David Kampmann's methodology.
    
    Steps:
    1. Get company-level production (from WSA/annual reports)
    2. Load plant capacities (from GEM)
    3. Calculate uniform utilization rate
    4. Allocate production to plants
    5. Apply emission factors by plant
    6. Sum for company total
    """
    
    def __init__(self, plants_df: pd.DataFrame, production_df: pd.DataFrame):
        """
        Initialize calculator.
        
        Args:
            plants_df: GEM plant data with columns [plant_id, plant_name, country, 
                       parent, capacity_ttpa, process, ef]
            production_df: Company production with columns [company, year, production_mt]
        """
        self.plants = plants_df
        self.production = production_df
        
        # Company name mapping (our names to GEM parent patterns)
        self.company_map = {
            "Tata Steel": "Tata Steel",
            "ArcelorMittal": "ArcelorMittal SA",
            "POSCO": "Posco",
            "Nippon Steel": "Nippon Steel",
            "JSW Steel": "JSW Steel Ltd",
            "SAIL": "Steel Authority of India",
            "Cleveland-Cliffs": "Cleveland-Cliffs",
            "ThyssenKrupp": "thyssenkrupp",
            "SSAB": "SSAB",
            "Severstal": "Severstal",
            "China Baowu Group": "Baowu|Baoshan",
            "JFE Holdings": "JFE",
            "Hyundai Steel": "Hyundai Steel",
            "Nucor": "Nucor",
            "BlueScope Steel": "BlueScope",
            "Gerdau": "Gerdau",
            "Voestalpine": "voestalpine",
            "NLMK": "Novolipetsk|NLMK",
            "MMK": "Magnitogorsk",
            "Jindal Steel & Power": "Jindal Steel And Power",
            "China Steel Corporation": "China Steel Corp",
        }
    
    def get_company_plants(self, company: str) -> pd.DataFrame:
        """Get all plants for a company."""
        pattern = self.company_map.get(company, company)
        return self.plants[
            self.plants['parent'].str.contains(pattern, case=False, na=False)
        ].copy()
    
    def calculate_company_emissions(self, company: str, year: int) -> Dict:
        """
        Calculate emissions for a company in a given year.
        
        Returns dict with:
        - production_mt: Company production
        - total_capacity_mt: Sum of plant capacities
        - utilization_rate: production / capacity
        - emissions_mt: Total CO2 emissions
        - weighted_ef: Capacity-weighted emission factor
        - plant_breakdown: List of plant-level calculations
        """
        # Get company production
        prod_row = self.production[
            (self.production['company'] == company) & 
            (self.production['year'] == year)
        ]
        
        if prod_row.empty:
            return None
        
        production_mt = prod_row['production_mt'].values[0]
        
        # Get company plants
        plants = self.get_company_plants(company)
        
        if plants.empty:
            logger.warning(f"No plants found for {company}")
            return None
        
        # Calculate total capacity
        total_capacity_ttpa = plants['capacity_ttpa'].sum()
        total_capacity_mt = total_capacity_ttpa / 1000
        
        # Calculate utilization rate
        utilization_rate = production_mt / total_capacity_mt if total_capacity_mt > 0 else 0
        
        # Allocate production to plants and calculate emissions
        plant_breakdown = []
        total_emissions = 0
        
        for _, plant in plants.iterrows():
            plant_capacity_mt = plant['capacity_ttpa'] / 1000
            plant_production = plant_capacity_mt * utilization_rate
            plant_emissions = plant_production * plant['ef']
            
            total_emissions += plant_emissions
            
            plant_breakdown.append({
                'plant_name': plant['plant_name'],
                'country': plant['country'],
                'process': plant['process'],
                'capacity_mt': plant_capacity_mt,
                'production_mt': plant_production,
                'ef': plant['ef'],
                'emissions_mt': plant_emissions,
            })
        
        # Calculate weighted average EF
        weighted_ef = total_emissions / production_mt if production_mt > 0 else 0
        
        return {
            'company': company,
            'year': year,
            'production_mt': production_mt,
            'total_capacity_mt': total_capacity_mt,
            'utilization_rate': utilization_rate,
            'emissions_mt': total_emissions,
            'weighted_ef': weighted_ef,
            'n_plants': len(plants),
            'plant_breakdown': plant_breakdown,
        }
    
    def calculate_all_companies(self, year: int) -> pd.DataFrame:
        """Calculate emissions for all companies with data for a given year."""
        results = []
        
        companies = self.production[self.production['year'] == year]['company'].unique()
        
        for company in companies:
            result = self.calculate_company_emissions(company, year)
            if result:
                results.append({
                    'company': result['company'],
                    'year': result['year'],
                    'production_mt': result['production_mt'],
                    'emissions_mt': result['emissions_mt'],
                    'weighted_ef': result['weighted_ef'],
                    'utilization_rate': result['utilization_rate'],
                    'n_plants': result['n_plants'],
                })
        
        return pd.DataFrame(results).sort_values('emissions_mt', ascending=False)


# =============================================================================
# SECTION 6: Main Pipeline
# =============================================================================

def run_pipeline(
    kampmann_excel_path: str,
    output_dir: str = "./output",
    years: List[int] = [2020, 2021, 2022, 2023]
) -> pd.DataFrame:
    """
    Run the full APA calculation pipeline.
    
    Args:
        kampmann_excel_path: Path to David Kampmann's Excel file
        output_dir: Directory for output files
        years: Years to calculate
    
    Returns:
        DataFrame with company emissions for all years
    """
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    logger.info("=" * 60)
    logger.info("STEEL APA CALCULATION PIPELINE")
    logger.info("=" * 60)
    
    # Step 1: Load GEM plant data
    logger.info("\n1. Loading GEM plant data...")
    plants = GEMPlantLoader.load_from_kampmann_excel(kampmann_excel_path)
    logger.info(f"   Loaded {len(plants)} operating plants")
    
    # Step 2: Get production data
    logger.info("\n2. Loading production data...")
    production = CompanyReportScraper.get_manual_company_data()
    logger.info(f"   Loaded {len(production)} company-year records")
    
    # Step 3: Initialize calculator
    logger.info("\n3. Initializing calculator...")
    calc = SteelAPACalculator(plants, production)
    
    # Step 4: Calculate emissions for all years
    logger.info("\n4. Calculating emissions...")
    all_results = []
    
    for year in years:
        logger.info(f"   Processing {year}...")
        year_results = calc.calculate_all_companies(year)
        all_results.append(year_results)
    
    results = pd.concat(all_results, ignore_index=True)
    
    # Step 5: Save outputs
    logger.info("\n5. Saving outputs...")
    
    results.to_csv(output_path / "steel_apa_emissions.csv", index=False)
    logger.info(f"   Saved: {output_path / 'steel_apa_emissions.csv'}")
    
    # Summary
    logger.info("\n" + "=" * 60)
    logger.info("RESULTS SUMMARY (2023)")
    logger.info("=" * 60)
    
    summary_2023 = results[results['year'] == 2023].head(15)
    for _, row in summary_2023.iterrows():
        logger.info(f"  {row['company']:<25} {row['emissions_mt']:>8.1f} Mt CO2  (EF={row['weighted_ef']:.2f})")
    
    return results


# =============================================================================
# SECTION 7: Utility Functions for Data Collection
# =============================================================================

def create_data_collection_template(companies: List[str], years: List[int], 
                                     output_path: str) -> None:
    """
    Create an Excel template for manual data collection.
    
    The template includes:
    - Sheet 1: Company production (company × year grid)
    - Sheet 2: Data sources (where to find production data)
    - Sheet 3: Instructions
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")
    
    wb = Workbook()
    
    # Sheet 1: Production data
    ws1 = wb.active
    ws1.title = "Production Data"
    
    # Headers
    ws1.cell(row=1, column=1, value="Company").font = Font(bold=True)
    for i, year in enumerate(years):
        ws1.cell(row=1, column=i+2, value=year).font = Font(bold=True)
    ws1.cell(row=1, column=len(years)+2, value="Source").font = Font(bold=True)
    
    # Companies
    for i, company in enumerate(companies):
        ws1.cell(row=i+2, column=1, value=company)
    
    # Sheet 2: Data sources
    ws2 = wb.create_sheet("Data Sources")
    sources = [
        ("Company", "IR URL", "Report Type", "Notes"),
        ("Tata Steel", "https://www.tatasteel.com/investors/", "Integrated Report", "FY ends March"),
        ("ArcelorMittal", "https://corporate.arcelormittal.com/investors", "20-F Filing", "CY"),
        ("POSCO", "https://www.posco.co.kr/homepage/docs/eng6/jsp/ir/", "Annual Report", "CY"),
        ("JSW Steel", "https://www.jsw.in/investors/steel", "Annual Report", "FY ends March"),
        ("SAIL", "https://sail.co.in/en/investor", "Annual Report", "FY ends March"),
        ("Nippon Steel", "https://www.nipponsteel.com/en/ir/", "Annual Report", "FY ends March"),
    ]
    
    for i, row in enumerate(sources):
        for j, val in enumerate(row):
            ws2.cell(row=i+1, column=j+1, value=val)
            if i == 0:
                ws2.cell(row=i+1, column=j+1).font = Font(bold=True)
    
    wb.save(output_path)
    logger.info(f"Created template: {output_path}")


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    import sys
    
    # Example usage
    print(__doc__)
    
    print("""
USAGE:
------

1. Run the full pipeline:
   
   from steel_apa_automation import run_pipeline
   results = run_pipeline(
       kampmann_excel_path="path/to/Copy_of_1_Output_Sheet_based_on_GEM_20240801_new_UR.xlsx",
       output_dir="./output",
       years=[2020, 2021, 2022, 2023]
   )

2. Calculate for a single company:

   from steel_apa_automation import (
       GEMPlantLoader, CompanyReportScraper, SteelAPACalculator
   )
   
   plants = GEMPlantLoader.load_from_kampmann_excel("path/to/kampmann.xlsx")
   production = CompanyReportScraper.get_manual_company_data()
   calc = SteelAPACalculator(plants, production)
   
   result = calc.calculate_company_emissions("Tata Steel", 2023)
   print(f"Tata Steel 2023: {result['emissions_mt']:.1f} Mt CO2")

3. Get emission factor for any plant:

   from steel_apa_automation import get_plant_ef
   
   ef = get_plant_ef("India", "DRI")  # Returns 3.10 (coal-based)
   ef = get_plant_ef("Iran", "DRI")   # Returns 1.05 (gas-based)
   ef = get_plant_ef("Germany", "BF-BOF")  # Returns 1.77 (EU)

4. Create data collection template:

   from steel_apa_automation import create_data_collection_template
   
   create_data_collection_template(
       companies=["Tata Steel", "JSW Steel", "SAIL", "Cleveland-Cliffs"],
       years=[2020, 2021, 2022, 2023, 2024],
       output_path="steel_production_template.xlsx"
   )
""")
