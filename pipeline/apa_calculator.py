"""
APA (Asset-based Planning Approach) Calculator for Steel Companies.

This module implements David Kampmann's "Asset-based Planning Approach"
methodology for estimating steel company CO2 emissions.

The APA approach:
1. Load plant data from GEM GIST database (capacity, technology, country)
2. Match plants to companies via Parent field
3. For a given (company, year, production_mt):
   - Build year-specific plant set (using start/close dates)
   - UR = production / total_capacity
   - Allocate production to each plant by capacity share
   - Apply country x technology emission factor per plant
   - Sum -> company total APA emissions

Emission factors sourced from:
  - Koolen & Vidovic (2022) JRC129297 for BF-BOF and EAF
  - IEA and academic literature for DRI factors

Data source: GEM Global Iron and Steel Tracker (GIST), December 2025 V1.
"""

import logging
import re
import sys
from functools import lru_cache
from pathlib import Path

import numpy as np
import pandas as pd

from .config import (
    EF_BF_BOF,
    GEM_FILE,
    GEM_STEEL_PLANTS_FILE,
    GEM_STEEL_IRON_UNITS_FILE,
    GEM_STEEL_STEEL_UNITS_FILE,
    KAMPMANN_ALD_FILE,
    KAMPMANN_EXCEL_FILE,
    PROCESSED_DATA_DIR,
    PROCESSED_STEEL_DIR,
)

logger = logging.getLogger(__name__)

# ============================================================================
# Emission factors
# ============================================================================

# Base emission factors — reference year ~2020
# Source: Koolen & Vidovic (2022) JRC129297 Table A4
# "Greenhouse gas intensities of the EU steel industry and its trading partners"
#
# Time-varying adjustment:
# Global BF-BOF intensity has remained largely flat over 2014-2024 (worldsteel
# Sustainability Indicators 2025). Country-specific annual time-series are not
# publicly available. We apply a small global improvement rate of 0.5%/year
# relative to the 2020 reference, based on:
#   - worldsteel global intensity trend: ~flat, minor improvements from best
#     practices adoption
#   - Global Efficiency Intelligence Steel Climate Impact 2022/2025: snapshot
#     comparisons show <5% change over a decade for most countries
#   - IEA Iron & Steel Technology Roadmap: 0.3-0.7% annual improvement assumed
# EAF scope 1 EFs are already very small (0.02-0.12 tCO2/t) so annual changes
# are negligible — kept static. DRI factors are technology-specific constants.
EF_REFERENCE_YEAR = 2020
EF_BF_BOF_ANNUAL_IMPROVEMENT = 0.005  # 0.5% annual reduction from reference year

# EAF emission factors by country (tCO2 per tonne crude steel)
# Values represent Scope 1 direct emissions only — static (changes negligible)
EF_EAF = {
    "Belarus": 0.07,
    "Brazil": 0.05,
    "China": 0.03,
    "EU": 0.04,
    "India": 0.07,
    "Japan": 0.04,
    "Russia": 0.07,
    "Serbia": 0.06,
    "South Africa": 0.12,
    "South Korea": 0.03,
    "Switzerland": 0.08,  # JRC Table A4 value (was 0.02 in original)
    "Taiwan": 0.02,
    "Thailand": 0.05,
    "Turkey": 0.04,
    "Ukraine": 0.04,
    "United Kingdom": 0.04,
    "United States": 0.04,
    # Kampmann-derived countries (use EU-comparable EAF where applicable)
    "Australia": 0.04,
    "New Zealand": 0.04,
    "Indonesia": 0.04,
    "Kazakhstan": 0.07,   # Russia-comparable
    "Canada": 0.04,       # US-comparable
    "Global": 0.051,
}

# DRI emission factors — technology-specific constants (static)
EF_DRI_COAL = 3.10   # Coal-based (India, China, South Africa)
EF_DRI_GAS = 1.05    # Gas-based (Iran, Middle East, Americas, etc.)
EF_H2_DRI = 0.04     # Hydrogen-based (future technology)

# Countries using COAL-based DRI (not gas)
# NZ Glenbrook uses ironsand direct-reduction in rotary kilns with coal
DRI_COAL_COUNTRIES = {"India", "China", "South Africa", "New Zealand"}

# Map plant country to EF lookup region
COUNTRY_TO_EF_REGION = {
    # Direct matches (countries with specific EF values from JRC129297)
    "Belarus": "Belarus",
    "Brazil": "Brazil",
    "China": "China",
    "India": "India",
    "Japan": "Japan",
    "Russia": "Russia",
    "Serbia": "Serbia",
    "South Africa": "South Africa",
    "South Korea": "South Korea",
    "Switzerland": "Switzerland",
    "Taiwan": "Taiwan",
    "Thailand": "Thailand",
    "Turkey": "Turkey",
    "Turkiye": "Turkey",
    "Ukraine": "Ukraine",
    "United Kingdom": "United Kingdom",
    "United States": "United States",
    # EU members
    "Germany": "EU",
    "France": "EU",
    "Netherlands": "EU",
    "Belgium": "EU",
    "Austria": "EU",
    "Spain": "EU",
    "Italy": "EU",
    "Sweden": "EU",
    "Finland": "EU",
    "Poland": "EU",
    "Czech Republic": "EU",
    "Czechia": "EU",
    "Luxembourg": "EU",
    "Romania": "EU",
    "Slovakia": "EU",
    "Hungary": "EU",
    "Croatia": "EU",
    "Greece": "EU",
    "Portugal": "EU",
    "Bulgaria": "EU",
    "Slovenia": "EU",
    "Latvia": "EU",
    "Estonia": "EU",
    "Lithuania": "EU",
    "Ireland": "EU",
    "Cyprus": "EU",
    "Malta": "EU",
    # Direct matches — Kampmann-derived EFs (now in EF_BF_BOF dict)
    "Australia": "Australia",
    "New Zealand": "New Zealand",
    "Indonesia": "Indonesia",
    "Kazakhstan": "Kazakhstan",
    "Canada": "Canada",
    # Proxies (countries without specific EF, mapped to similar regions)
    "Mexico": "Brazil",
    "Vietnam": "China",
    "Malaysia": "China",
    "Philippines": "India",
    "Bangladesh": "India",
    "Iran": "Turkey",
    "Egypt": "Turkey",
    "Argentina": "Brazil",
    "Colombia": "Brazil",
    "Peru": "Brazil",
    "Chile": "Brazil",
    "Norway": "EU",
    "Saudi Arabia": "Turkey",
    "United Arab Emirates": "Turkey",
    "Qatar": "Turkey",
    "Oman": "Turkey",
    "Bahrain": "Turkey",
    "Kuwait": "Turkey",
    "Algeria": "Turkey",
    "Libya": "Turkey",
    "Tunisia": "Turkey",
    "Morocco": "Turkey",
    "Pakistan": "India",
    "Myanmar": "India",
    "Trinidad and Tobago": "Brazil",
    "Venezuela": "Brazil",
    "Ecuador": "Brazil",
    "Bolivia": "Brazil",
    "Uruguay": "Brazil",
    "Paraguay": "Brazil",
}

# ============================================================================
# Company matching patterns (GEM Parent / Owner field)
# ============================================================================

COMPANY_GEM_PATTERNS = {
    # Original 14 companies
    "ArcelorMittal": r"ArcelorMittal",
    "Tata Steel": r"Tata Steel",
    "POSCO Holdings": r"Posco|POSCO",
    "Nippon Steel": r"Nippon Steel",
    "JSW Steel": r"JSW Steel|JSW Ispat",
    "ThyssenKrupp": r"[Tt]hyssenkrupp|ThyssenKrupp",
    "SSAB": r"SSAB",
    "Severstal": r"Severstal",
    "Baoshan Iron & Steel": r"Baoshan|Baowu",
    "BlueScope Steel": r"BlueScope",
    "China Steel": r"China Steel|Dragon Steel",
    "Nucor": r"Nucor",
    "Gerdau": r"Gerdau",
    "JFE Holdings": r"JFE",
    # Expanded companies
    "US Steel": r"U\.?S\.?\s*Steel|United States Steel",
    "Hyundai Steel": r"Hyundai Steel",
    "Cleveland-Cliffs": r"Cleveland.Cliffs|Cliffs Natural|AK Steel",
    "Kobe Steel": r"Kobe Steel|KOBELCO",
    "voestalpine": r"voestalpine",
    "SAIL": r"Steel Authority|SAIL",
    "Steel Dynamics": r"Steel Dynamics|SDI\b",
    "Salzgitter": r"Salzgitter",
    "Ternium": r"Ternium",
    "NLMK": r"NLMK|Novolipetsk",
    "Evraz": r"Evraz",
    "Liberty Steel": r"Liberty Steel|GFG Alliance",
}

# Plant name-based patterns for companies whose GEM Parent changed due to
# acquisitions (e.g. US Steel → Nippon Steel in Dec 2024).
# These are matched against plant_name as a fallback when parent match fails.
COMPANY_PLANT_NAME_PATTERNS = {
    "US Steel": r"U\.?S\.?\s*Steel",
}

# Ownership transfers: plants that changed hands during our analysis period.
# Format: (acquirer, target_plant_name_pattern, year_acquired)
# Before year_acquired: target company owns plants (parent match fails, use name)
# After year_acquired: acquirer owns plants (parent match succeeds)
# The target company's plants should be EXCLUDED from the acquirer for years < year_acquired.
OWNERSHIP_TRANSFERS = [
    {
        "acquirer": "Nippon Steel",
        "target": "US Steel",
        "target_plant_pattern": r"U\.?S\.?\s*Steel",
        "year_acquired": 2024,  # Nippon Steel acquisition closed Dec 2024
    },
]


# ============================================================================
# Emission factor lookup
# ============================================================================

def get_ef_region(country: str) -> str:
    """Map a plant country to its emission factor region."""
    return COUNTRY_TO_EF_REGION.get(country, "Global")


def get_plant_ef(country: str, process: str, year: int | None = None) -> float:
    """Get emission factor for a plant based on country and process type.

    Matches David Kampmann's methodology exactly:
    - BF-BOF: country-specific from Definitions sheet, with year adjustment
    - EAF: country-specific (grid carbon intensity dependent), static
    - DRI: Coal-based (3.10) for India/China/SA, Gas-based (1.05) elsewhere
    - H2-DRI: 0.04

    For BF-BOF, applies a global annual improvement factor relative to the
    EF_REFERENCE_YEAR (2020). The adjustment is small (~0.5%/year) and reflects
    the gradual efficiency improvements observed in worldsteel trend data.

    Args:
        country: Plant country from GSPT (e.g. "Germany", "India")
        process: Classified process type ("BF-BOF", "EAF", "DRI", "H2-DRI")
        year: Optional year for time-adjusted BF-BOF EF. If None, uses
              reference year (no adjustment).

    Returns:
        Emission factor in tCO2 per tonne crude steel.
    """
    region = get_ef_region(country)
    process_lower = process.lower()

    if "dri" in process_lower and "h2" not in process_lower:
        # DRI -- check if coal or gas based
        if country in DRI_COAL_COUNTRIES:
            return EF_DRI_COAL
        return EF_DRI_GAS
    elif "eaf" in process_lower or "electric" in process_lower or "scrap" in process_lower:
        return EF_EAF.get(region, EF_EAF["Global"])
    elif "h2" in process_lower or "hydrogen" in process_lower:
        return EF_H2_DRI
    else:
        # Default: BF-BOF with year adjustment
        base_ef = EF_BF_BOF.get(region, EF_BF_BOF["Global"])
        if year is not None:
            years_from_ref = year - EF_REFERENCE_YEAR
            # Apply compound improvement: EF decreases over time
            adjustment = (1 - EF_BF_BOF_ANNUAL_IMPROVEMENT) ** years_from_ref
            return base_ef * adjustment
        return base_ef


# ============================================================================
# Plant data loading — GEM GIST December 2025 format
# ============================================================================

_plants_cache: pd.DataFrame | None = None


def _parse_year(val) -> float:
    """Extract a year from a date field (could be int, float, datetime, or string)."""
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float)):
        v = float(val)
        if 1900 <= v <= 2100:
            return v
        return np.nan
    s = str(val).strip()
    # Try to extract a 4-digit year
    m = re.search(r"(19|20)\d{2}", s)
    if m:
        return float(m.group())
    return np.nan


def _determine_process(row: pd.Series) -> str:
    """Determine plant technology from capacity columns and equipment text.

    Classification logic (capacity-based):
    1. If plant has BOTH BF and DRI capacity, classify by the dominant one:
       - BF >= DRI → "BF-BOF" (e.g. POSCO Pohang: BF=12,900, DRI=300)
       - DRI > BF → "DRI" (e.g. ArcelorMittal Hazira: DRI=7,830, BF=2,040)
    2. If only BF > 0 → "BF-BOF"
    3. If only DRI > 0 → "DRI"
    4. If only EAF > 0 → "EAF"
    5. Fallback to equipment text, then default "BF-BOF"
    """
    dri = row.get("dri_capacity", 0)
    bf = row.get("bf_capacity", 0)
    eaf = row.get("eaf_capacity", 0)

    has_bf = pd.notna(bf) and bf > 0
    has_dri = pd.notna(dri) and dri > 0
    has_eaf = pd.notna(eaf) and eaf > 0

    # When both BF and DRI are present, classify by the dominant process
    if has_bf and has_dri:
        return "BF-BOF" if bf >= dri else "DRI"
    if has_bf:
        return "BF-BOF"
    if has_dri:
        return "DRI"
    if has_eaf:
        return "EAF"

    # Fallback to equipment text
    equip = str(row.get("main_equipment", "")).lower()
    if re.search(r"bf|blast.?furnace|bof|basic.?oxygen", equip):
        return "BF-BOF"
    if re.search(r"dri|sponge|direct.?red", equip):
        return "DRI"
    if re.search(r"eaf|electric|scrap", equip):
        return "EAF"
    if re.search(r"h2|hydrogen", equip):
        return "H2-DRI"

    return "BF-BOF"


def load_all_gem_plants(gem_path: Path | None = None) -> pd.DataFrame:
    """Load ALL steel plants from GEM GIST December 2025 multi-sheet format.

    Reads the Plant-level data file, joining Plant data with Plant capacities
    to get per-plant capacity and status. Includes all statuses (operating,
    retired, mothballed, announced, construction) for year-specific filtering.

    Results are cached after first load (Excel reading is slow).

    Args:
        gem_path: Path to the GEM plant-level Excel file.
                  Defaults to config.GEM_STEEL_PLANTS_FILE.

    Returns:
        DataFrame with columns: plant_id, plant_name, country, parent,
        status, capacity_ttpa, bf_capacity, eaf_capacity, dri_capacity,
        main_equipment, process, ef, start_year, close_year
    """
    global _plants_cache
    if _plants_cache is not None:
        return _plants_cache

    if gem_path is None:
        gem_path = GEM_STEEL_PLANTS_FILE

    if not gem_path.exists():
        logger.error(f"GEM plant file not found: {gem_path}")
        return pd.DataFrame()

    logger.info(f"Loading GEM GIST plant data from {gem_path.name}...")

    # --- Sheet 1: Plant data (identity, ownership, start date) ---
    plant_data = pd.read_excel(gem_path, sheet_name="Plant data")

    # Find columns flexibly
    plant_col_map = {}
    desired = {
        "Plant ID": "plant_id",
        "Plant name (English)": "plant_name",
        "Country": "country",
        "Parent": "parent",
        "Owner": "owner",
        "Start": "start_raw",
        "Main production equipment": "main_equipment",
    }
    for orig, new in desired.items():
        matches = [c for c in plant_data.columns if orig.lower() in c.lower()]
        if matches:
            plant_col_map[matches[0]] = new

    plants = plant_data[list(plant_col_map.keys())].rename(columns=plant_col_map).copy()

    # Parse start year
    if "start_raw" in plants.columns:
        plants["start_year"] = plants["start_raw"].apply(_parse_year)
        plants = plants.drop(columns=["start_raw"])
    else:
        plants["start_year"] = np.nan

    # --- Sheet 2: Plant capacities and status ---
    caps = pd.read_excel(gem_path, sheet_name="Plant capacities and status")

    cap_col_map = {}
    desired_caps = {
        "Plant ID": "plant_id",
        "Status": "status",
        "Start": "cap_start_raw",
        "Nominal crude steel capacity (ttpa)": "capacity_ttpa",
        "Nominal BF capacity (ttpa)": "bf_capacity",
        "Nominal BOF steel capacity (ttpa)": "bof_capacity",
        "Nominal EAF steel capacity (ttpa)": "eaf_capacity",
        "Nominal DRI capacity (ttpa)": "dri_capacity",
    }
    for orig, new in desired_caps.items():
        matches = [c for c in caps.columns if orig.lower() in c.lower()]
        if matches:
            cap_col_map[matches[0]] = new

    caps_df = caps[list(cap_col_map.keys())].rename(columns=cap_col_map).copy()

    # Convert capacity columns to numeric
    for col in ["capacity_ttpa", "bf_capacity", "bof_capacity", "eaf_capacity", "dri_capacity"]:
        if col in caps_df.columns:
            caps_df[col] = pd.to_numeric(caps_df[col], errors="coerce")

    # Parse capacity-sheet start year (unit-level start date)
    # This is distinct from the plant-level start date in Sheet 1.
    # For construction entries, this gives the expected commissioning year
    # (e.g. Glenbrook EAF = 2026), not the plant founding year (1969).
    if "cap_start_raw" in caps_df.columns:
        caps_df["cap_start_year"] = caps_df["cap_start_raw"].apply(_parse_year)
        caps_df = caps_df.drop(columns=["cap_start_raw"])
    else:
        caps_df["cap_start_year"] = np.nan

    # The capacities sheet has multiple rows per plant (one per status phase).
    # We need to handle this: for year-specific filtering, we want each status
    # phase as a separate record. But for a given year, only certain statuses
    # are "active". So we aggregate capacity per plant_id x status.
    #
    # Key statuses:
    #   - "operating": currently active
    #   - "retired": no longer active (need close date)
    #   - "mothballed": temporarily inactive
    #   - "operating pre-retirement": still active but announced closure
    #   - "construction": coming online soon
    #   - "announced": not yet built
    #   - "cancelled": will not be built
    #
    # For capacity, we sum across all rows for a given plant_id + status.
    # A plant can have both "operating" and "retired" capacity tranches.

    # Aggregate capacity by plant_id and status
    cap_agg = caps_df.groupby(["plant_id", "status"]).agg({
        "capacity_ttpa": "sum",
        "bf_capacity": "sum",
        "eaf_capacity": "sum",
        "dri_capacity": "sum",
        "cap_start_year": "min",  # earliest start date within group
    }).reset_index()

    # Filter to statuses that represent real or historical capacity
    active_statuses = {
        "operating", "retired", "mothballed",
        "operating pre-retirement", "construction",
    }
    cap_active = cap_agg[cap_agg["status"].str.lower().isin(active_statuses)].copy()

    # Filter to positive capacity
    cap_active = cap_active[
        cap_active["capacity_ttpa"].notna() & (cap_active["capacity_ttpa"] > 0)
    ].copy()

    # Join with plant identity data
    result = cap_active.merge(plants, on="plant_id", how="left")

    # Use parent field; if empty, fall back to owner
    if "owner" in result.columns:
        result["parent"] = result["parent"].fillna(result["owner"])
        result = result.drop(columns=["owner"])

    # Parse close dates for retired/mothballed plants
    # The GIST Dec 2025 doesn't have an explicit close_date column in the
    # plant-level file. We infer:
    # - "operating" / "operating pre-retirement" / "construction": close_year = NaN (still active)
    # - "retired" / "mothballed": These plants have stopped; we don't have an
    #   exact close date from this file, so we'll leave close_year = NaN and
    #   handle them by status alone.
    result["close_year"] = np.nan

    # Prefer capacity-sheet start year (unit-level) over plant-level start year
    # for construction entries. The plant-level start_year is the plant founding
    # date (e.g. Glenbrook 1969), but for new construction units at existing
    # plants, the capacity sheet has the actual commissioning year (e.g. 2026).
    if "cap_start_year" in result.columns:
        # For construction entries: always use capacity-level start year if available
        construction_cap_mask = (
            result["status"].str.lower() == "construction"
        ) & result["cap_start_year"].notna()
        result.loc[construction_cap_mask, "start_year"] = result.loc[
            construction_cap_mask, "cap_start_year"
        ]
        # For other statuses: use cap_start_year only if plant-level is missing
        other_mask = (
            result["status"].str.lower() != "construction"
        ) & result["start_year"].isna() & result["cap_start_year"].notna()
        result.loc[other_mask, "start_year"] = result.loc[other_mask, "cap_start_year"]
        result = result.drop(columns=["cap_start_year"])

    # For operating plants with unknown start_year, assume pre-2000 (well before 2014)
    status_lower = result["status"].str.lower()
    operating_mask = status_lower.isin({"operating", "operating pre-retirement"})
    result.loc[operating_mask & result["start_year"].isna(), "start_year"] = 2000.0

    # For retired/mothballed plants with unknown start_year, also assume pre-2000
    retired_mask = status_lower.isin({"retired", "mothballed"})
    result.loc[retired_mask & result["start_year"].isna(), "start_year"] = 2000.0

    # For construction plants with unknown start_year, assume 2025
    construction_mask = status_lower == "construction"
    result.loc[construction_mask & result["start_year"].isna(), "start_year"] = 2025.0

    # Determine process type
    result["process"] = result.apply(_determine_process, axis=1)

    # Reclassify integrated DRI-EAF plants: when the same plant_id has both
    # DRI and EAF entries, the EAF is part of the integrated DRI steelmaking
    # process (DRI ironmaking → EAF steelmaking), NOT a separate scrap-fed
    # mini-mill. The EAF in such plants should use the DRI emission factor,
    # not the low scrap-EAF factor (0.04).
    # Example: BlueScope Glenbrook NZ has DRI (ironsand reduction) + EAF
    # entries — the EAF refines DRI iron, not recycled scrap.
    dri_plant_ids = set(result.loc[result["process"] == "DRI", "plant_id"])
    integrated_eaf_mask = (
        result["process"] == "EAF"
    ) & result["plant_id"].isin(dri_plant_ids)
    if integrated_eaf_mask.any():
        n_reclass = integrated_eaf_mask.sum()
        plants_affected = result.loc[integrated_eaf_mask, "plant_name"].unique()
        logger.info(
            f"Reclassified {n_reclass} integrated DRI-EAF entries as DRI: "
            f"{', '.join(str(p) for p in plants_affected)}"
        )
        result.loc[integrated_eaf_mask, "process"] = "DRI"

    # Assign emission factor
    result["ef"] = result.apply(
        lambda r: get_plant_ef(str(r.get("country", "")), str(r["process"])), axis=1
    )

    n_operating = (status_lower == "operating").sum()
    n_retired = retired_mask.sum()
    n_pre_ret = (status_lower == "operating pre-retirement").sum()
    n_construction = construction_mask.sum()
    logger.info(
        f"Loaded {len(result)} plant-status entries: "
        f"{n_operating} operating, {n_pre_ret} pre-retirement, "
        f"{n_retired} retired/mothballed, {n_construction} construction"
    )

    _plants_cache = result.reset_index(drop=True)
    return _plants_cache


def get_plants_for_year(all_plants: pd.DataFrame, year: int) -> pd.DataFrame:
    """Filter plants to those active in a given year.

    Rules:
    - "operating" / "operating pre-retirement": include if start_year <= year
    - "retired" / "mothballed": include if start_year <= year
      (We include all retired plants for historical years because we don't have
       exact close dates from the GIST plant file. The Dec 2025 GIST marks
       plants as "retired" but doesn't provide the retirement year in the
       plant-level data. Since most retirements in our 2014-2024 window are
       recent, including them gives better historical capacity estimates than
       excluding them entirely. Future improvement: cross-reference with the
       iron unit data for BF stop dates.)
    - "construction": include if start_year <= year (it came online)
    - "announced" / "cancelled": always exclude

    Args:
        all_plants: Full plant DataFrame from load_all_gem_plants()
        year: The year to filter for

    Returns:
        Filtered DataFrame of plants active in the given year.
    """
    if all_plants.empty:
        return all_plants

    status_lower = all_plants["status"].str.lower()

    # Operating / pre-retirement: include if started by this year
    operating_mask = (
        status_lower.isin({"operating", "operating pre-retirement"})
        & (all_plants["start_year"] <= year)
    )

    # Construction: include if it came online by this year
    construction_mask = (
        (status_lower == "construction")
        & (all_plants["start_year"] <= year)
    )

    # Retired/mothballed: include for all historical years
    # (conservative — assumes they were operating during 2014-2024)
    # Only include if they started before or in this year
    retired_mask = (
        status_lower.isin({"retired", "mothballed"})
        & (all_plants["start_year"] <= year)
    )

    combined = operating_mask | construction_mask | retired_mask
    return all_plants[combined].copy()


# Backwards compatibility wrapper
def load_gem_plants(gem_path: Path | None = None) -> pd.DataFrame:
    """Load GEM plants (backwards compatible — returns only operating plants).

    For new code, use load_all_gem_plants() + get_plants_for_year() instead.
    """
    all_plants = load_all_gem_plants(gem_path)
    if all_plants.empty:
        return all_plants
    # Return only currently operating plants (like the old behavior)
    status_lower = all_plants["status"].str.lower()
    return all_plants[status_lower.isin({"operating", "operating pre-retirement"})].copy()


# ============================================================================
# Company matching
# ============================================================================

def get_company_plants(plants_df: pd.DataFrame, company: str, year: int | None = None) -> pd.DataFrame:
    """Filter plants to those belonging to a specific company.

    Uses regex pattern matching against the Parent field. If no match,
    falls back to plant_name patterns (for companies whose GEM ownership
    changed due to acquisitions, e.g. US Steel → Nippon Steel).

    Handles ownership transfers: if company X acquired company Y's plants
    in year Z, then for year < Z, X's plants exclude Y's plants, and Y's
    plants are matched by plant_name instead.

    Args:
        plants_df: Plant DataFrame (full or year-filtered)
        company: Canonical company name (key in COMPANY_GEM_PATTERNS)
        year: Optional year for ownership-aware filtering

    Returns:
        Filtered DataFrame of matching plants, or empty DataFrame.
    """
    pattern = COMPANY_GEM_PATTERNS.get(company, company)
    mask = plants_df["parent"].str.contains(pattern, case=False, na=False)
    result = plants_df[mask].copy()

    # Handle ownership transfers
    if year is not None and not result.empty:
        for transfer in OWNERSHIP_TRANSFERS:
            if transfer["acquirer"] == company and year < transfer["year_acquired"]:
                # Before acquisition: exclude target's plants from acquirer
                exclude_pattern = transfer["target_plant_pattern"]
                exclude_mask = result["plant_name"].str.contains(
                    exclude_pattern, case=False, na=False
                )
                n_excluded = exclude_mask.sum()
                if n_excluded > 0:
                    result = result[~exclude_mask].copy()
                    logger.debug(
                        f"Excluded {n_excluded} {transfer['target']} plants from "
                        f"{company} for {year} (pre-acquisition)"
                    )

    # Fallback: match by plant name (for ownership changes)
    if result.empty and company in COMPANY_PLANT_NAME_PATTERNS:
        name_pattern = COMPANY_PLANT_NAME_PATTERNS[company]
        name_mask = plants_df["plant_name"].str.contains(name_pattern, case=False, na=False)
        result = plants_df[name_mask].copy()
        if not result.empty:
            logger.debug(f"Matched {len(result)} plants for '{company}' via plant_name fallback")

    if result.empty:
        logger.debug(f"No plants found for '{company}' (pattern: {pattern})")

    return result


# ============================================================================
# Core APA calculation
# ============================================================================

def calculate_company_emissions(
    plants_df: pd.DataFrame,
    company: str,
    production_mt: float,
    year: int | None = None,
    country_production: dict[str, float] | None = None,
) -> dict | None:
    """Calculate APA emissions for a single company-year.

    Supports two production allocation modes:

    **Country-level allocation** (when ``country_production`` is provided):
    For each country with a known production figure, plants in that country
    share the stated production proportionally by capacity.  This replicates
    Kampmann's methodology where country-level production is sourced from
    annual reports and sustainability reports.

    **Uniform UR fallback** (when ``country_production`` is None):
    A single company-wide utilization rate is applied to all plants.

    Args:
        plants_df: Plant DataFrame (should be year-filtered)
        company: Canonical company name
        production_mt: Company crude steel production in million tonnes
        year: Optional year for ownership-aware plant filtering and
              time-adjusted BF-BOF emission factors.
        country_production: Optional dict mapping country name to production
              in Mt.  When provided, production is allocated at the country
              level rather than uniformly.

    Returns:
        Dict with emissions_mt, weighted_ef, utilization_rate, n_plants,
        plant_breakdown. Returns None if no plants found.
    """
    company_plants = get_company_plants(plants_df, company, year=year)

    if company_plants.empty:
        return None

    total_capacity_mt = company_plants["capacity_ttpa"].sum() / 1000.0
    if total_capacity_mt <= 0:
        return None

    ur = production_mt / total_capacity_mt

    # Allocate production and compute emissions per plant
    breakdown = company_plants[
        ["plant_id", "plant_name", "country", "process",
         "capacity_ttpa", "ef"]
    ].copy()

    # Apply year-specific EF adjustment for BF-BOF plants
    if year is not None:
        breakdown["ef"] = breakdown.apply(
            lambda r: get_plant_ef(str(r["country"]), str(r["process"]), year=year),
            axis=1,
        )

    breakdown["capacity_mt"] = breakdown["capacity_ttpa"] / 1000.0

    if country_production:
        # --- Country-level production allocation ---
        # For each country with known production, compute a country-specific UR
        # and allocate production to plants in that country proportionally.
        # Countries without explicit production data get the residual production
        # allocated via a uniform UR across their plants.
        allocated = pd.Series(0.0, index=breakdown.index)
        remaining_production = production_mt
        unmatched_mask = pd.Series(True, index=breakdown.index)

        for country, prod_mt_country in country_production.items():
            mask = breakdown["country"] == country
            country_cap = breakdown.loc[mask, "capacity_mt"].sum()
            if country_cap > 0 and mask.any():
                country_ur = prod_mt_country / country_cap
                allocated[mask] = breakdown.loc[mask, "capacity_mt"] * country_ur
                remaining_production -= prod_mt_country
                unmatched_mask[mask] = False

        # Allocate remaining production to unmatched countries uniformly
        if remaining_production > 0 and unmatched_mask.any():
            unmatched_cap = breakdown.loc[unmatched_mask, "capacity_mt"].sum()
            if unmatched_cap > 0:
                residual_ur = remaining_production / unmatched_cap
                allocated[unmatched_mask] = (
                    breakdown.loc[unmatched_mask, "capacity_mt"] * residual_ur
                )

        breakdown["allocated_production_mt"] = allocated
    else:
        # --- Uniform UR allocation (original approach) ---
        breakdown["allocated_production_mt"] = breakdown["capacity_mt"] * ur

    breakdown["emissions_mt"] = breakdown["allocated_production_mt"] * breakdown["ef"]

    total_emissions = breakdown["emissions_mt"].sum()
    weighted_ef = total_emissions / production_mt if production_mt > 0 else 0.0

    return {
        "company": company,
        "production_mt": production_mt,
        "total_capacity_mt": round(total_capacity_mt, 3),
        "utilization_rate": round(ur, 4),
        "emissions_mt": round(total_emissions, 3),
        "weighted_ef": round(weighted_ef, 3),
        "n_plants": len(company_plants),
        "plant_breakdown": breakdown,
    }


# ============================================================================
# Production data loading
# ============================================================================

def load_bau_production(gem_path: Path | None = None) -> pd.DataFrame:
    """Extract 'Reported crude steel output' from BAU sheets in the Kampmann workbook.

    Iterates over sheets matching *_BAU, finds the row containing
    'Reported crude steel output', reads year columns.

    Production values are in kt in the workbook; this function converts to Mt.

    Returns:
        DataFrame with columns: company, year, production_mt, production_source
    """
    kampmann_path = KAMPMANN_EXCEL_FILE
    if gem_path is not None and gem_path.exists():
        # Legacy: if gem_path points to the Kampmann Excel, use it
        kampmann_path = gem_path

    if not kampmann_path.exists():
        logger.warning(f"Kampmann file not found for BAU production: {kampmann_path}")
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    import openpyxl
    wb = openpyxl.load_workbook(kampmann_path, data_only=True, read_only=True)

    bau_sheets = [s for s in wb.sheetnames if s.strip().endswith("_BAU")]
    records = []

    for sheet_name in bau_sheets:
        company_name = sheet_name.strip().replace("_BAU", "").strip()
        # Map to canonical names
        company_canonical = _bau_company_to_canonical(company_name)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            continue

        # Find year header row (row index 0 or 1 typically has years 2020, 2021, ...)
        year_row = None
        year_col_map = {}  # col_index -> year
        for ri, row in enumerate(rows[:5]):
            for ci, val in enumerate(row):
                if isinstance(val, (int, float)) and 2015 <= val <= 2050:
                    year_row = ri
                    break
            if year_row is not None:
                break

        if year_row is None:
            continue

        # Build column-to-year map
        for ci, val in enumerate(rows[year_row]):
            if isinstance(val, (int, float)) and 2015 <= val <= 2050:
                year_col_map[ci] = int(val)

        # Find "Reported crude steel output" row
        for row in rows:
            label = str(row[0] if row[0] else "") + str(row[1] if len(row) > 1 and row[1] else "")
            if "reported crude steel output" in label.lower():
                for ci, year in year_col_map.items():
                    if ci < len(row) and row[ci] is not None:
                        try:
                            val_kt = float(row[ci])
                            if val_kt > 0:
                                records.append({
                                    "company": company_canonical,
                                    "year": year,
                                    "production_mt": round(val_kt / 1000, 3),
                                    "production_source": "bau_reported",
                                })
                        except (ValueError, TypeError):
                            pass
                break

    wb.close()

    df = pd.DataFrame(records)
    if not df.empty:
        logger.info(f"BAU production: {len(df)} records, "
                    f"{df['company'].nunique()} companies, "
                    f"years {df['year'].min()}-{df['year'].max()}")
    return df


def _bau_company_to_canonical(bau_name: str) -> str:
    """Map BAU sheet company names to canonical names."""
    mapping = {
        "ArcelorMittal": "ArcelorMittal",
        "Baoshan": "Baoshan Iron & Steel",
        "BlueScope": "BlueScope Steel",
        "China Steel": "China Steel",
        "POSCO": "POSCO Holdings",
        "ThyssenKrupp": "ThyssenKrupp",
        "Severstal": "Severstal",
        "Nippon Steel": "Nippon Steel",
        "SSAB": "SSAB",
        "Tata Steel": "Tata Steel",
    }
    return mapping.get(bau_name, bau_name)


def load_annual_report_production() -> pd.DataFrame:
    """Load production data from annual report extractions.

    Reads steel_all_extracted.csv, filters to production_mt metric,
    excludes known bad extractions.

    Returns:
        DataFrame with columns: company, year, production_mt, production_source
    """
    extracted_file = PROCESSED_STEEL_DIR / "steel_all_extracted.csv"
    if not extracted_file.exists():
        logger.warning(f"Extracted data not found: {extracted_file}")
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    df = pd.read_csv(extracted_file)
    prod = df[df["metric"] == "production_mt"].copy()

    if prod.empty:
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    # Exclude known bad extractions
    bad = [
        ("JFE Holdings", 2015),  # 100 Mt = cumulative milestone
        ("JFE Holdings", 2016),  # 100 Mt = cumulative milestone
        ("JFE Holdings", 2017),  # 3.04 Mt = unit error
        ("Nippon Steel", 2024),  # 52 Mt = capacity not production
    ]
    for company, year in bad:
        prod = prod[~((prod["company"] == company) & (prod["year"] == year))]

    result = prod[["company", "year", "value"]].rename(
        columns={"value": "production_mt"}
    ).copy()
    result["production_source"] = "annual_report"

    if not result.empty:
        logger.info(f"Annual report production: {len(result)} records, "
                    f"{result['company'].nunique()} companies")
    return result


def load_curated_production() -> pd.DataFrame:
    """Load production data from hand-curated steel_production_from_reports.csv.

    This file contains manually verified production figures sourced from
    annual reports, 10-K filings, sustainability reports, and Q4 earnings
    press releases. ArcelorMittal values use consolidated crude steel
    production (from Q4 results), not WorldSteel figures that include JVs.

    Returns:
        DataFrame with columns: company, year, production_mt, production_source
    """
    curated_file = PROCESSED_STEEL_DIR / "steel_production_from_reports.csv"
    if not curated_file.exists():
        logger.warning(f"Curated production file not found: {curated_file}")
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    df = pd.read_csv(curated_file)

    # Map company names to canonical names
    name_map = {
        "ArcelorMittal": "ArcelorMittal",
        "Nippon Steel": "Nippon Steel",
        "POSCO": "POSCO Holdings",
        "Tata Steel": "Tata Steel",
        "JSW Steel": "JSW Steel",
        "JFE Holdings": "JFE Holdings",
        "Nucor": "Nucor",
        "ThyssenKrupp": "ThyssenKrupp",
        "SSAB": "SSAB",
        "Gerdau": "Gerdau",
        "BlueScope": "BlueScope Steel",
        "Severstal": "Severstal",
        "China Steel": "China Steel",
        "Cleveland-Cliffs": "Cleveland-Cliffs",
        "Hyundai Steel": "Hyundai Steel",
        "voestalpine": "voestalpine",
        "US Steel": "US Steel",
        "Kobe Steel": "Kobe Steel",
        "SAIL": "SAIL",
        "Steel Dynamics": "Steel Dynamics",
        "Salzgitter": "Salzgitter",
        "Ternium": "Ternium",
        "NLMK": "NLMK",
        "Evraz": "Evraz",
        "Liberty Steel": "Liberty Steel",
    }
    df["company"] = df["company"].map(name_map).fillna(df["company"])

    result = df[["company", "year", "production_mt"]].copy()
    result["production_source"] = "curated_reports"

    if not result.empty:
        logger.info(f"Curated production: {len(result)} records, "
                    f"{result['company'].nunique()} companies")
    return result


def load_ald_production() -> pd.DataFrame:
    """Load production data from Kampmann ALD CSV.

    Values labeled 'MtSteel' are actually in kt -- divide by 1000.
    Only BAU scenario, years 2020-2025.

    Returns:
        DataFrame with columns: company, year, production_mt, production_source
    """
    if not KAMPMANN_ALD_FILE.exists():
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    df = pd.read_csv(KAMPMANN_ALD_FILE)

    # Filter to BAU Production rows
    prod_cols = [c for c in df.columns if "Production" in c and "BAU" in c]
    if not prod_cols:
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    prod_col = prod_cols[0]

    records = []
    for _, row in df.iterrows():
        year = row.get("Year")
        if not isinstance(year, (int, float)) or year < 2020 or year > 2025:
            continue

        company_raw = row.get("Company Name", "")
        company = _bau_company_to_canonical(company_raw)
        val = row.get(prod_col)

        if pd.notna(val) and val > 0:
            records.append({
                "company": company,
                "year": int(year),
                "production_mt": round(float(val) / 1000, 3),  # kt to Mt
                "production_source": "kampmann_ald",
            })

    result = pd.DataFrame(records)
    # Aggregate by company-year (ALD has per-country rows)
    if not result.empty:
        result = (
            result.groupby(["company", "year", "production_source"])
            .agg({"production_mt": "sum"})
            .reset_index()
        )
        result["production_mt"] = result["production_mt"].round(3)
        logger.info(f"ALD production: {len(result)} records, "
                    f"{result['company'].nunique()} companies")

    return result


def load_gem_plant_production() -> pd.DataFrame:
    """Load production data aggregated from GEM GIST plant-level production.

    The GIST Plant Production sheet has per-plant crude steel output for
    2019-2024 (coverage varies by plant and year). We aggregate by company
    using the same parent-matching patterns as get_company_plants().

    This is the LOWEST priority source — only used as a fallback when no
    other production data is available for a company-year.

    Returns:
        DataFrame with columns: company, year, production_mt, production_source
    """
    if not GEM_STEEL_PLANTS_FILE.exists():
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    try:
        prod = pd.read_excel(GEM_STEEL_PLANTS_FILE, sheet_name="Plant production")
        plant_data = pd.read_excel(GEM_STEEL_PLANTS_FILE, sheet_name="Plant data")
    except Exception as e:
        logger.warning(f"Failed to read GEM plant production: {e}")
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    # Only crude steel production
    crude = prod[prod["Type of production"] == "Crude steel production (ttpa)"].copy()

    # Join with parent from plant data
    parent_map = plant_data[["Plant ID", "Parent"]].drop_duplicates("Plant ID")
    crude = crude.merge(parent_map, on="Plant ID", how="left")

    # Get plant names for fallback matching
    name_col = [c for c in plant_data.columns if "name" in c.lower() and "english" in c.lower()]
    if name_col:
        name_map = plant_data[["Plant ID", name_col[0]]].drop_duplicates("Plant ID")
        name_map = name_map.rename(columns={name_col[0]: "plant_name"})
        crude = crude.merge(name_map, on="Plant ID", how="left")

    year_cols = [c for c in crude.columns if isinstance(c, int) and 2014 <= c <= 2030]
    for yr in year_cols:
        crude[yr] = pd.to_numeric(crude[yr], errors="coerce")

    records = []
    for company, pattern in COMPANY_GEM_PATTERNS.items():
        mask = crude["Parent"].str.contains(pattern, case=False, na=False)

        # Plant name fallback
        if company in COMPANY_PLANT_NAME_PATTERNS and "plant_name" in crude.columns:
            name_pat = COMPANY_PLANT_NAME_PATTERNS[company]
            name_mask = crude["plant_name"].str.contains(name_pat, case=False, na=False)
            mask = mask | name_mask

        company_rows = crude[mask]
        if company_rows.empty:
            continue

        # Find best-covered year (most plants reporting) to set coverage floor
        plants_per_year = {}
        for yr in year_cols:
            n_reporting = company_rows[yr].notna().sum()
            plants_per_year[yr] = n_reporting
        max_plants = max(plants_per_year.values()) if plants_per_year else 0

        for yr in year_cols:
            total_ttpa = company_rows[yr].sum()
            n_reporting = plants_per_year.get(yr, 0)
            if total_ttpa > 0 and max_plants > 0:
                # Skip if <50% of plants reported vs best year (partial data)
                coverage = n_reporting / max_plants
                if coverage < 0.5:
                    logger.debug(
                        f"Skipping GEM production for {company} {yr}: "
                        f"{n_reporting}/{max_plants} plants reporting ({coverage:.0%})"
                    )
                    continue
                records.append({
                    "company": company,
                    "year": int(yr),
                    "production_mt": round(total_ttpa / 1000, 3),
                    "production_source": "gem_plant_level",
                })

    result = pd.DataFrame(records)
    if not result.empty:
        logger.info(f"GEM plant production: {len(result)} records, "
                    f"{result['company'].nunique()} companies, "
                    f"years {result['year'].min()}-{result['year'].max()}")
    return result


def load_production_data(gem_path: Path | None = None) -> pd.DataFrame:
    """Aggregate production data from all sources with priority.

    Follows Kampmann et al. (2024) production sourcing hierarchy:
      "Where available, we retrieved P(disclosed) and CAP(disclosed) for
      production plants from mandatory financial reporting (e.g., annual
      reports, investor presentations), voluntary financial reporting
      (CDP reports), or other company sources such as company websites
      (in order of priority). Where company disclosure on PR(disclosed)
      was missing, we took the total crude steel production reported by
      the World Steel Association."

    Our implementation:
      0. bau_reported — Kampmann's own verified figures from his Excel workbook
         (Row 121 "Reported crude steel output"), originally sourced from
         annual reports by Kampmann. Available for 10 companies, 2020-2023.
      1. annual_report — extracted from downloaded annual report PDFs by our
         automated extraction pipeline.
      2. curated_reports — hand-curated production from annual reports, 10-K
         filings, sustainability reports, and Q4 earnings press releases.
         Verified against Kampmann's figures where overlap exists.
      3. kampmann_ald — Kampmann ALD CSV aggregated production (backup).
      4. gem_plant_level — GEM GIST plant-level crude steel output (lowest
         priority, subject to 50% coverage floor filter).

    For each (company, year), keeps the highest-priority available value.

    Returns:
        DataFrame with columns: company, year, production_mt, production_source
    """
    bau = load_bau_production(gem_path)
    ar = load_annual_report_production()
    curated = load_curated_production()
    ald = load_ald_production()
    gem_prod = load_gem_plant_production()

    # Assign priority (lower = higher priority)
    # All top-3 sources are ultimately from annual reports (Kampmann's approach).
    # WSA is NOT used as a primary source — only as validation reference.
    bau["_priority"] = 0       # Kampmann's own annual-report-sourced figures
    ar["_priority"] = 1        # Our annual report extraction
    curated["_priority"] = 2   # Our hand-curated from annual reports
    ald["_priority"] = 3       # Kampmann ALD aggregated (backup)
    gem_prod["_priority"] = 4  # GEM plant-level (lowest priority)

    all_prod = pd.concat([bau, ar, curated, ald, gem_prod], ignore_index=True)

    if all_prod.empty:
        logger.warning("No production data found from any source")
        return pd.DataFrame(columns=["company", "year", "production_mt", "production_source"])

    # For each (company, year), keep highest priority
    all_prod = all_prod.sort_values("_priority")
    result = all_prod.drop_duplicates(subset=["company", "year"], keep="first")
    result = result.drop(columns=["_priority"]).reset_index(drop=True)

    logger.info(f"Production data: {len(result)} company-year pairs, "
                f"{result['company'].nunique()} companies, "
                f"years {result['year'].min()}-{result['year'].max()}")

    # Show per-company summary
    for company in sorted(result["company"].unique()):
        comp = result[result["company"] == company]
        years = sorted(comp["year"].tolist())
        sources = comp["production_source"].unique().tolist()
        logger.info(f"  {company}: {len(years)} years ({min(years)}-{max(years)}), "
                    f"sources: {sources}")

    return result


# ============================================================================
# Country-level production allocation
# ============================================================================

# Map Kampmann ALD company names to our canonical names
_KAMPMANN_TO_CANONICAL = {
    "POSCO": "POSCO Holdings",
    # All others match exactly
}


def generate_country_production(gem_path: Path | None = None) -> pd.DataFrame:
    """Generate country-level production CSV with 4-tier priority hierarchy.

    Sources (highest priority first):
    0. **Curated from annual reports** — hand-verified country production from
       sustainability/annual reports (e.g., SSAB emissions-intensity derived).
    1. **Kampmann ALD year-specific (2020-2024)** — country-level production from
       Kampmann's academic dataset, capturing structural shifts (Ukraine collapse,
       Italy closure, US expansion) even though proportions may be modelled.
    2. **Kampmann 2020 proportions scaled** — for pre-2020 years (2014-2019),
       scale 2020 country proportions by company production ratio.
    3. **GEM capacity shares** — for non-Kampmann companies, allocate production
       proportionally by plant capacity in each country.

    A (company, year) pair is only filled by a lower-priority source if no
    higher-priority source has already covered it.

    Returns DataFrame and saves to steel_country_production.csv.
    """
    from .config import (
        KAMPMANN_ALD_FILE, PROCESSED_STEEL_DIR,
        CURATED_COUNTRY_PRODUCTION_FILE,
    )

    rows: list[dict] = []
    covered_pairs: set[tuple[str, int]] = set()  # (company, year) already covered

    # --- Priority 0: Curated from annual reports ---
    curated_count = 0
    if CURATED_COUNTRY_PRODUCTION_FILE.exists():
        curated = pd.read_csv(CURATED_COUNTRY_PRODUCTION_FILE)
        for _, r in curated.iterrows():
            company = str(r["company"])
            year = int(r["year"])
            rows.append({
                "company": company,
                "year": year,
                "country": r["country"],
                "production_kt": float(r["production_kt"]),
                "source": f"curated_{r['source']}",
            })
            covered_pairs.add((company, year))
            curated_count += 1
        logger.info(f"Country production: loaded {curated_count} curated rows "
                    f"from annual reports")

    # --- Priority 1: Kampmann ALD year-specific (2020-2024) ---
    kampmann_companies: set[str] = set()
    kampmann_2020: dict[str, dict[str, float]] = {}  # canonical_name -> {country: kt}
    kampmann_ald_count = 0

    if KAMPMANN_ALD_FILE.exists():
        k = pd.read_csv(KAMPMANN_ALD_FILE)
        k_prod = k[
            (k["Variable"] == "Production (BAU)") &
            (k["Year"].between(2020, 2024)) &
            (k["Value"] > 0)
        ]
        for _, r in k_prod.iterrows():
            raw_name = r["Company Name"]
            canonical = _KAMPMANN_TO_CANONICAL.get(raw_name, raw_name)
            kampmann_companies.add(canonical)
            year = int(r["Year"])

            # Always store 2020 data for pre-2020 scaling fallback
            if year == 2020:
                if canonical not in kampmann_2020:
                    kampmann_2020[canonical] = {}
                kampmann_2020[canonical][r["Country"]] = r["Value"]  # in kt

            # Only add if not already covered by curated data
            if (canonical, year) not in covered_pairs:
                rows.append({
                    "company": canonical,
                    "year": year,
                    "country": r["Country"],
                    "production_kt": r["Value"],
                    "source": "kampmann_ald",
                })
                kampmann_ald_count += 1

        # Mark all Kampmann ALD (company, year) pairs as covered
        for _, r in k_prod.iterrows():
            canonical = _KAMPMANN_TO_CANONICAL.get(r["Company Name"], r["Company Name"])
            covered_pairs.add((canonical, int(r["Year"])))

        logger.info(f"Country production: loaded {kampmann_ald_count} Kampmann ALD rows "
                    f"for {len(kampmann_companies)} companies (2020-2024)")

    # --- Priority 2: Scale Kampmann 2020 proportions for pre-2020 years ---
    company_production = load_production_data(gem_path)
    scaled_count = 0

    for canonical in kampmann_companies:
        if canonical not in kampmann_2020:
            continue
        country_shares_2020 = kampmann_2020[canonical]
        total_2020_kt = sum(country_shares_2020.values())
        if total_2020_kt <= 0:
            continue

        # Get company production for scaling
        comp_prod = company_production[company_production["company"] == canonical]
        prod_2020_row = comp_prod[comp_prod["year"] == 2020]
        prod_2020_mt = (
            prod_2020_row.iloc[0]["production_mt"] if len(prod_2020_row) > 0
            else total_2020_kt / 1000.0
        )

        for _, prod_row in comp_prod.iterrows():
            year = int(prod_row["year"])
            if (canonical, year) in covered_pairs:
                continue  # Already covered by curated or ALD data

            # Scale 2020 country proportions by production ratio
            scale = prod_row["production_mt"] / prod_2020_mt if prod_2020_mt > 0 else 1.0
            for country, prod_kt_2020 in country_shares_2020.items():
                rows.append({
                    "company": canonical,
                    "year": year,
                    "country": country,
                    "production_kt": prod_kt_2020 * scale,
                    "source": "kampmann_scaled",
                })
                scaled_count += 1
            covered_pairs.add((canonical, year))

    if scaled_count:
        logger.info(f"Country production: {scaled_count} scaled rows (pre-2020)")

    # --- Priority 3: GEM capacity shares for non-Kampmann companies ---
    all_plants = load_all_gem_plants(gem_path)
    gem_count = 0
    if not all_plants.empty:
        non_k = company_production[
            ~company_production["company"].isin(kampmann_companies)
        ]
        for _, prod_row in non_k.iterrows():
            company = prod_row["company"]
            year = int(prod_row["year"])
            prod_mt = prod_row["production_mt"]

            if (company, year) in covered_pairs:
                continue

            year_plants = get_plants_for_year(all_plants, year)
            comp_plants = get_company_plants(year_plants, company, year=year)
            if comp_plants.empty:
                continue

            total_cap = comp_plants["capacity_ttpa"].sum()
            if total_cap <= 0:
                continue

            # Group by country
            for country, grp in comp_plants.groupby("country"):
                country_cap = grp["capacity_ttpa"].sum()
                share = country_cap / total_cap
                rows.append({
                    "company": company,
                    "year": year,
                    "country": country,
                    "production_kt": prod_mt * 1000.0 * share,  # Mt → kt
                    "source": "gem_capacity_share",
                })
                gem_count += 1
            covered_pairs.add((company, year))

    if gem_count:
        logger.info(f"Country production: {gem_count} GEM capacity share rows")

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["company", "year", "country"]).reset_index(drop=True)

    # Save
    PROCESSED_STEEL_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_STEEL_DIR / "steel_country_production.csv"
    df.to_csv(out_path, index=False)
    logger.info(f"Saved country production: {out_path} "
                f"({len(df)} rows, {df['company'].nunique()} companies, "
                f"sources: {df['source'].value_counts().to_dict() if not df.empty else {}})")

    return df


def load_country_production() -> dict[tuple[str, int], dict[str, float]]:
    """Load country-level production data from CSV.

    Returns:
        Dict mapping (company, year) → {country_name: production_mt}.
        Empty dict if file not found.
    """
    from .config import PROCESSED_STEEL_DIR

    path = PROCESSED_STEEL_DIR / "steel_country_production.csv"
    if not path.exists():
        logger.info("No country-level production file found, using uniform UR")
        return {}

    df = pd.read_csv(path)
    result: dict[tuple[str, int], dict[str, float]] = {}
    for (company, year), grp in df.groupby(["company", "year"]):
        result[(str(company), int(year))] = {
            row["country"]: row["production_kt"] / 1000.0  # kt → Mt
            for _, row in grp.iterrows()
            if row["production_kt"] > 0
        }

    logger.info(f"Country production: {len(result)} (company, year) pairs loaded")
    return result


def estimate_production_from_capacity(
    all_plants: pd.DataFrame,
    production: pd.DataFrame,
    fill_years: range | list[int] = range(2014, 2020),
) -> pd.DataFrame:
    """Estimate production for company-years missing data using GEM capacity.

    For each company that has at least some years with both known production
    and known GEM capacity, compute a historical utilization rate (UR) as the
    average UR from the earliest 3 such years.  Then for each requested
    fill_year where production is missing, estimate:

        production_mt = company_capacity_mt(year) × avg_UR

    Args:
        all_plants: Full GEM plant DataFrame (all years, from load_all_gem_plants).
        production: Existing production DataFrame with (company, year, production_mt,
            production_source) — output of load_production_data().
        fill_years: Years to attempt to fill (default 2014–2019).

    Returns:
        DataFrame with same schema as *production* containing only the newly
        estimated rows (production_source="capacity_estimate").
    """
    if all_plants.empty or production.empty:
        return pd.DataFrame(columns=production.columns)

    existing_keys = set(zip(production["company"], production["year"].astype(int)))
    companies = production["company"].unique()

    rows: list[dict] = []
    for company in companies:
        comp_prod = production[production["company"] == company].copy()
        comp_prod["year"] = comp_prod["year"].astype(int)

        # For each known-production year, compute GEM capacity → UR
        ur_samples: list[tuple[int, float]] = []
        for _, pr in comp_prod.iterrows():
            yr = int(pr["year"])
            year_plants = get_plants_for_year(all_plants, yr)
            comp_plants = get_company_plants(year_plants, company, year=yr)
            cap_mt = comp_plants["capacity_ttpa"].sum() / 1000.0
            if cap_mt > 0:
                ur = pr["production_mt"] / cap_mt
                if 0.1 <= ur <= 1.5:  # sanity bounds
                    ur_samples.append((yr, ur))

        if not ur_samples:
            continue

        # Average UR from earliest 3 overlap years
        ur_samples.sort(key=lambda x: x[0])
        earliest = ur_samples[:3]
        avg_ur = sum(u for _, u in earliest) / len(earliest)
        yr_range = f"{earliest[0][0]}-{earliest[-1][0]}" if len(earliest) > 1 else str(earliest[0][0])

        for yr in fill_years:
            if (company, yr) in existing_keys:
                continue

            year_plants = get_plants_for_year(all_plants, yr)
            comp_plants = get_company_plants(year_plants, company, year=yr)
            cap_mt = comp_plants["capacity_ttpa"].sum() / 1000.0
            if cap_mt <= 0:
                continue

            est_prod = cap_mt * avg_ur
            rows.append({
                "company": company,
                "year": yr,
                "production_mt": round(est_prod, 3),
                "production_source": "capacity_estimate",
            })
            logger.info(
                f"  Estimated {company} {yr}: {est_prod:.2f} Mt "
                f"(cap={cap_mt:.1f} Mt × UR={avg_ur:.2f} from {yr_range})"
            )

    result = pd.DataFrame(rows, columns=production.columns if rows else
                          ["company", "year", "production_mt", "production_source"])
    if not result.empty:
        logger.info(
            f"Capacity-estimated production: {len(result)} company-year rows "
            f"for {result['company'].nunique()} companies"
        )
    return result


# ============================================================================
# Full APA pipeline
# ============================================================================

def run_apa_all(gem_path: Path | None = None) -> pd.DataFrame:
    """Run APA calculation for all company-year combinations with data.

    Uses year-specific plant filtering: for each (company, year), builds
    the plant set that was active in that year using start/close dates.

    1. Loads ALL GEM plants (all statuses)
    2. Loads production data from all sources
    3. For each (company, year): filters plants to that year, calculates emissions

    Returns:
        DataFrame with company, year, production_mt, emissions_mt, weighted_ef,
        utilization_rate, n_plants, production_source
    """
    all_plants = load_all_gem_plants(gem_path)
    if all_plants.empty:
        return pd.DataFrame()

    production = load_production_data()
    if production.empty:
        return pd.DataFrame()

    # Fill gaps for pre-2020 years using capacity × historical UR
    estimated = estimate_production_from_capacity(
        all_plants, production, fill_years=range(2014, 2020),
    )
    if not estimated.empty:
        production = pd.concat([production, estimated], ignore_index=True)
        logger.info(f"Production after capacity estimates: {len(production)} total rows")

    # Load country-level production for accurate allocation
    country_prod_map = load_country_production()
    if country_prod_map:
        logger.info(f"Using country-level production for {len(country_prod_map)} company-years")

    results = []
    for _, row in production.iterrows():
        company = row["company"]
        year = int(row["year"])
        prod_mt = row["production_mt"]

        # Get year-specific plant set
        year_plants = get_plants_for_year(all_plants, year)

        # Get country-level production if available
        country_prod = country_prod_map.get((company, year), None)

        result = calculate_company_emissions(
            year_plants, company, prod_mt, year=year,
            country_production=country_prod,
        )
        if result is not None:
            # Skip if utilization rate > 1.5 -- indicates incomplete plant coverage
            # (Relaxed from 1.0 to 1.5 to accommodate minor GEM capacity gaps)
            if result["utilization_rate"] > 1.5:
                logger.warning(
                    f"Skipping {company} {year}: UR={result['utilization_rate']:.2f} > 1.5 "
                    f"({result['n_plants']} plants, {result['total_capacity_mt']:.1f} Mt capacity "
                    f"vs {prod_mt:.1f} Mt production)"
                )
                continue
            results.append({
                "company": company,
                "year": year,
                "production_mt": prod_mt,
                "emissions_mt": result["emissions_mt"],
                "weighted_ef": result["weighted_ef"],
                "utilization_rate": result["utilization_rate"],
                "n_plants": result["n_plants"],
                "total_capacity_mt": result["total_capacity_mt"],
                "production_source": row["production_source"],
            })

    df = pd.DataFrame(results)
    if not df.empty:
        df = df.sort_values(["company", "year"]).reset_index(drop=True)
        logger.info(f"APA results: {len(df)} company-year calculations, "
                    f"{df['company'].nunique()} companies")

    return df


# ============================================================================
# Integration bridge for integrate.py
# ============================================================================

def load_apa_source() -> pd.DataFrame:
    """Format APA results for the multi-source integration pipeline.

    Returns data in the same row format as load_annual_reports(),
    load_climate_trace(), and load_kampmann_ald() in integrate.py.

    Each company-year produces two rows:
      - production_mt (echoes input production)
      - emissions_mt_co2 (APA-calculated emissions)
    """
    apa = run_apa_all()
    if apa.empty:
        logger.warning("No APA results to integrate")
        return pd.DataFrame()

    rows = []
    for _, r in apa.iterrows():
        detail = (f"{r['n_plants']} plants, "
                  f"UR={r['utilization_rate']:.2f}, "
                  f"wEF={r['weighted_ef']:.3f}")

        # Production row
        rows.append({
            "company": r["company"],
            "year": int(r["year"]),
            "metric": "production_mt",
            "value": r["production_mt"],
            "unit": "Mt",
            "source": "apa",
            "source_detail": detail,
            "extraction_method": "asset_level_model",
            "confidence_raw": "modeled" if r["production_source"] == "capacity_estimate" else "reported",
            "source_page": "",
            "notes": f"Production from {r['production_source']}",
        })

        # Emissions row
        rows.append({
            "company": r["company"],
            "year": int(r["year"]),
            "metric": "emissions_mt_co2",
            "value": r["emissions_mt"],
            "unit": "Mt CO2",
            "source": "apa",
            "source_detail": detail,
            "extraction_method": "asset_level_model",
            "confidence_raw": "modeled",
            "source_page": "",
            "notes": (f"APA: {r['production_mt']:.1f} Mt production "
                      f"from {r['production_source']}, "
                      f"capacity={r['total_capacity_mt']:.1f} Mt"),
        })

    result = pd.DataFrame(rows)
    logger.info(f"APA source: {len(result)} records for integration")
    return result


# ============================================================================
# Standalone execution
# ============================================================================

def _print_results_table(df: pd.DataFrame):
    """Print a formatted results table."""
    print("\n" + "=" * 90)
    print(f"{'Company':<25} {'Year':>5} {'Prod(Mt)':>9} {'Emis(Mt)':>9} "
          f"{'wEF':>6} {'UR':>6} {'Plants':>6} {'Source':<15}")
    print("-" * 90)

    for _, r in df.iterrows():
        print(f"{r['company']:<25} {int(r['year']):>5} "
              f"{r['production_mt']:>9.2f} {r['emissions_mt']:>9.2f} "
              f"{r['weighted_ef']:>6.3f} {r['utilization_rate']:>6.2f} "
              f"{r['n_plants']:>6} {r['production_source']:<15}")

    print("=" * 90)


def main():
    """Run APA calculator standalone with full output."""
    from .config import OUTPUTS_COMPANY_STEEL

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )

    print("=" * 60)
    print("STEEL APA CALCULATOR (GIST Dec 2025)")
    print("=" * 60)

    # Step 0: Generate ownership mapping for transparency
    print("\n--- Generating ownership mapping ---")
    try:
        from .ownership_mapping import generate_ownership_mapping
        generate_ownership_mapping()
    except Exception as e:
        logger.warning(f"Ownership mapping generation failed (non-fatal): {e}")

    # Step 1: Generate country-level production CSV
    print("\n--- Generating country-level production data ---")
    generate_country_production()

    # Step 2: Run APA with country-level allocation
    results = run_apa_all()

    if results.empty:
        print("No results computed.")
        return

    _print_results_table(results)

    # Save to standard output location
    OUTPUTS_COMPANY_STEEL.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUTS_COMPANY_STEEL / "steel_apa_emissions.csv"
    results.to_csv(out_path, index=False)
    print(f"\nSaved: {out_path} ({len(results)} rows)")

    # Summary statistics
    print(f"\nTotal: {len(results)} company-year calculations")
    print(f"Companies: {results['company'].nunique()}")
    print(f"Year range: {results['year'].min()}-{results['year'].max()}")
    print(f"\nBy company (latest year):")

    latest = results.sort_values("year").drop_duplicates("company", keep="last")
    latest = latest.sort_values("emissions_mt", ascending=False)
    for _, r in latest.iterrows():
        print(f"  {r['company']:<25} {int(r['year'])} "
              f"| {r['emissions_mt']:>7.2f} Mt CO2 "
              f"| EF={r['weighted_ef']:.3f} "
              f"| {r['n_plants']} plants")


if __name__ == "__main__":
    main()
